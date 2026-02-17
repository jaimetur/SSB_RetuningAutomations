# -*- coding: utf-8 -*-

import os
import time
import re
import shutil
import tempfile
from typing import List, Tuple, Optional, Dict
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

from src.utils.utils_io import find_log_files, read_text_file, to_long_path, pretty_path
from src.utils.utils_parsing import SUMMARY_RE, find_all_subnetwork_headers, extract_mo_from_subnetwork_line, parse_table_slice_from_subnetwork, parse_log_lines, find_subnetwork_header_index, extract_mo_name_from_previous_line, cap_rows
from src.utils.utils_excel import sanitize_sheet_name, unique_sheet_name, color_summary_tabs, apply_alternating_category_row_fills, style_headers_autofilter_and_autofit, style_headers_autofilter_and_autofit_xlsxwriter
from src.utils.utils_sorting import natural_logfile_key
from src.utils.utils_pivot import safe_pivot_count, safe_crosstab_count, apply_frequency_column_filter
from src.utils.utils_dataframe import concat_or_empty
from src.utils.utils_datetime import log_phase_timer, format_duration_hms
from src.modules.Common.correction_commands_exporter import export_all_sheets_with_correction_commands, export_external_and_termpoint_commands
from .ca_summary_excel import build_summary_audit
from .ca_summary_ppt import generate_ppt_summary


class ConfigurationAudit:
    """
    Generates an Excel in input_dir with one sheet per *.log / *.logs / *.txt file.
    (Functionality kept, extended with SummaryAudit sheet and PPT summary.)

    ARFCN-related parameters (N77, etc.) are now configurable via __init__:
      - new_arfcn            → main "new" NR / LTE ARFCN (e.g. 648672)
      - old_arfcn            → main "old" NR / LTE ARFCN (e.g. 647328)
      - allowed_n77_ssb      → allowed SSB values for N77 (e.g. {648672, 653952})
      - allowed_n77_arfcn    → allowed ARFCN values for N77 sectors
    """

    SUMMARY_RE = SUMMARY_RE  # keep class reference

    def __init__(
        self,
        n77_ssb_pre: int,
        n77_ssb_post: int,
        allowed_n77_ssb_pre: Optional[set[int]] = None,
        allowed_n77_arfcn_pre: Optional[set[int]] = None,
        allowed_n77_ssb_post: Optional[set[int]] = None,
        allowed_n77_arfcn_post: Optional[set[int]] = None,
        n77b_ssb_arfcn: Optional[int] = None
    ):
        """
        Initialize ConfigurationAudit with ARFCN-related parameters.

        All values are converted to integers/sets of integers internally to make checks robust.
        """
        # Core ARFCN values
        self.N77_SSB_PRE: int = int(n77_ssb_pre)
        self.N77_SSB_POST: int = int(n77_ssb_post)
        self.N77B_SSB: int = int(n77b_ssb_arfcn)

        # Allowed SSB (Pre) values for N77 cells (e.g. {648672, 653952})
        if allowed_n77_ssb_pre is None:
            self.ALLOWED_N77_SSB_PRE = set()
        else:
            self.ALLOWED_N77_SSB_PRE = {int(v) for v in allowed_n77_ssb_pre}

        # Allowed ARFCN (Pre) values for N77 sectors (e.g. {654652, 655324, 655984, 656656})
        if allowed_n77_arfcn_pre is None:
            self.ALLOWED_N77_ARFCN_PRE = set()
        else:
            self.ALLOWED_N77_ARFCN_PRE = {int(v) for v in allowed_n77_arfcn_pre}

        # Allowed SSB (Post) values for N77 cells (e.g. {648672, 653952})
        if allowed_n77_ssb_post is None:
            self.ALLOWED_N77_SSB_POST = set()
        else:
            self.ALLOWED_N77_SSB_POST = {int(v) for v in allowed_n77_ssb_post}

        # Allowed ARFCN (Post) values for N77 sectors (e.g. {654652, 655324, 655984, 656656})
        if allowed_n77_arfcn_post is None:
            self.ALLOWED_N77_ARFCN_POST = set()
        else:
            self.ALLOWED_N77_ARFCN_POST = {int(v) for v in allowed_n77_arfcn_post}

    # =====================================================================
    #                            PUBLIC API
    # =====================================================================
    def run(
            self,
            input_dir: str,
            module_name: Optional[str] = "",
            versioned_suffix: Optional[str] = None,
            tables_order: Optional[List[str]] = None,  # optional sheet ordering
            filter_frequencies: Optional[List[str]] = None,  # substrings to filter pivot columns
            output_dir: Optional[str] = None,  # <<< NEW: optional dedicated output folder
            profiles_audit: bool = True,  # <<< NEW: enable Profiles audit logic
            frequency_audit: bool = True,  # <<< NEW: show/hide NR/LTE frequency audits in SummaryAudit (NRFrequency, GUtranSyncSignalFrequency)
            show_phase_starts: bool = False,  # <<< NEW: show only START lines (no END lines)
            show_phase_timings: bool = True,  # <<< NEW: show timings as [INFO]
            slow_file_seconds_threshold: float = 10.0,  # <<< NEW: report per-file parsing when a file is slow
            slow_sheet_seconds_threshold: float = 10.0, # <<< NEW: report per-sheet when a sheet is slow
            source_zip_path: Optional[str] = None,  # <<< NEW: original ZIP path containing the logs (if input was a ZIP)
            extracted_root: Optional[str] = None,  # <<< NEW: temp extraction root folder (used to rebuild ZIP-based LogPath)
            export_correction_cmd: bool = True,
            correction_cmd_folder_name: str = "Correction_Cmd_CA",
            fast_excel_export: bool = False,
            fast_excel_autofit_rows: int = 50,
            fast_excel_autofit_max_width: int = 60
    ) -> str:

        """
        Main entry point: creates an Excel file with one sheet per detected table.
        Sheets are ordered according to TABLES_ORDER if provided; otherwise,
        they are sorted in a natural order by filename (Data_Collection.txt, Data_Collection(1).txt, ...).

        If 'filter_frequencies' is provided, the three added summary sheets will keep only
        those pivot *columns* whose header contains any of the provided substrings
        (case-insensitive). 'NodeId' and 'Total' are always kept.

        In addition, a 'SummaryAudit' sheet is created with high-level checks
        across the parsed tables, and a PowerPoint (.pptx) summary is generated
        with a textual bullet-style overview per category.

        Optional:
          - If profiles_audit=True, profiles tables will be collected and checked for old/new SSB replica consistency.
        """
        prefix = f"{module_name} " if module_name else ""

        _LEVEL_PREFIX_RE = re.compile(r"^\s*\[(INFO|DEBUG|WARNING|WARN|ERROR)\]\s*", re.IGNORECASE)

        def _ensure_level_prefix(message: str, level: str) -> str:
            """
            Ensure message starts with a single level prefix like '[INFO] '.
            If it already starts with any '[LEVEL]', keep it to avoid duplicates like '[INFO] [INFO] ...'.
            """
            msg = "" if message is None else str(message)
            if _LEVEL_PREFIX_RE.match(msg):
                return msg.strip()
            lvl = (level or "INFO").upper().strip()
            if lvl == "WARN":
                lvl = "WARNING"
            return f"[{lvl}] {msg}".strip()

        def _log(level: str, message: str) -> None:
            print(f"{prefix}{_ensure_level_prefix(message, level)}")

        def _log_info(message: str) -> None:
            _log("INFO", message)

        def _log_warn(message: str) -> None:
            _log("WARNING", message)

        def _make_temp_xlsx_path(final_xlsx_path: str) -> Tuple[str, str]:
            """
            Create a temp dir and a temp xlsx path. Prefer system temp to avoid OneDrive sync during write.
            Returns (tmp_dir, tmp_xlsx_path).
            """
            tmp_dir = tempfile.mkdtemp(prefix="SSB_RA_")
            tmp_name = os.path.basename(final_xlsx_path)
            tmp_xlsx = os.path.join(tmp_dir, tmp_name)
            return tmp_dir, tmp_xlsx

        def _move_into_place(tmp_path: str, final_path: str) -> None:
            """
            Move temp file into final destination.
            - Try atomic replace if possible.
            - Fallback to shutil.move (handles cross-device).
            """
            try:
                os.replace(tmp_path, final_path)
            except Exception:
                shutil.move(tmp_path, final_path)

        overall_start = time.perf_counter()

        with log_phase_timer("ConfigurationAudit", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
            # --- Normalize filters ---
            with log_phase_timer("PHASE 0.1: Normalize filters", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                freq_filters = [str(f).strip() for f in (filter_frequencies or []) if str(f).strip()]

            # --- Validate the input directory ---
            with log_phase_timer("PHASE 0.2: Validate input directory", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                if not os.path.isdir(input_dir):
                    raise NotADirectoryError(f"Invalid directory: {input_dir}")

            # <<< NEW: decide the base output folder and ensure it exists >>>
            # If output_dir is provided, all generated files (Excel/PPT) will be written there.
            # Otherwise, legacy behavior is kept and files are created under input_dir.
            with log_phase_timer("PHASE 0.3: Prepare output directory", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                base_output_dir = output_dir or input_dir
                base_output_dir_long = to_long_path(base_output_dir)
                os.makedirs(base_output_dir_long, exist_ok=True)

            # --- Detect log/txt files ---
            with log_phase_timer("PHASE 0.4: Detect log/txt files", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                log_files = find_log_files(input_dir)
                if not log_files:
                    _log_info(f"No log/txt files found in: '{pretty_path(input_dir)}'")
                    return ""

            # --- Natural sorting of files (handles '(1)', '(2)', '(10)', etc.) ---
            with log_phase_timer("PHASE 0.5: Sort files (natural order)", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                sorted_files = sorted(log_files, key=natural_logfile_key)
                file_rank: Dict[str, int] = {os.path.basename(p): i for i, p in enumerate(sorted_files)}

            # --- Build MO (table) ranking if TABLES_ORDER is provided ---
            with log_phase_timer("PHASE 0.6: Build MO rank (optional TABLES_ORDER)", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                mo_rank: Dict[str, int] = {}
                if tables_order:
                    mo_rank = {name: i for i, name in enumerate(tables_order)}

            # --- Prepare Excel output path ---
            # prefix = "ProfilesAudit" if profiles_audit else "ConfigurationAudit"
            prefix_name = "ConfigurationAudit"
            excel_path = os.path.join(base_output_dir_long, f"{prefix_name}_{versioned_suffix}.xlsx")
            excel_path_long = to_long_path(excel_path)

            table_entries: List[Dict[str, object]] = []

            # --- Keep a per-file index to preserve order of multiple tables inside same file ---
            per_file_table_idx: Dict[str, int] = {}

            # =====================================================================
            #                PHASE 1: Parse all log/txt files
            # =====================================================================
            _log_info("PHASE 1: Parse all log/txt files (this phase can take some time)...")

            with log_phase_timer("PHASE 1: Parse all log/txt files", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                file_counter = 0
                for i, path in enumerate(log_files, start=1):
                    file_counter += 1

                    base_filename = os.path.basename(path)
                    lines, encoding_used = read_text_file(path)

                    header_indices = find_all_subnetwork_headers(lines)

                    if not header_indices:
                        table_start = time.perf_counter()

                        header_idx = find_subnetwork_header_index(lines, self.SUMMARY_RE)
                        df, note = parse_log_lines(lines, self.SUMMARY_RE, forced_header_idx=header_idx)
                        mo_name_prev = extract_mo_name_from_previous_line(lines, header_idx)

                        if encoding_used:
                            note = (note + " | " if note else "") + f"encoding={encoding_used}"
                        df, note = cap_rows(df, note)

                        table_elapsed = time.perf_counter() - table_start

                        mo_name_for_log = mo_name_prev if mo_name_prev else ""
                        if header_indices and header_indices[0] < len(lines):
                            mo_from_subnet = extract_mo_from_subnetwork_line(lines[header_indices[0]])
                            if mo_from_subnet:
                                mo_name_for_log = mo_from_subnet
                        if not mo_name_for_log:
                            mo_name_for_log = "MO NOT FOUND"

                        if show_phase_timings:
                            tag = "[SLOW]" if table_elapsed >= float(slow_file_seconds_threshold) else ""
                            _log_info(f"PHASE 1: Parse all log/txt files - MO parse {file_counter:>3}: '{mo_name_for_log}' (File: {base_filename}) --> took {table_elapsed:.3f}s {tag}")

                        idx_in_file = per_file_table_idx.get(base_filename, 0)
                        per_file_table_idx[base_filename] = idx_in_file + 1

                        table_entries.append({"df": df, "sheet_candidate": mo_name_for_log, "log_file": base_filename, "tables_in_log": 1, "note": note or "", "idx_in_file": idx_in_file})

                    else:
                        tables_in_log = len(header_indices)
                        header_indices.append(len(lines))  # add sentinel index

                        file_start = time.perf_counter()
                        any_slow = False
                        mo_names: List[str] = []

                        for ix in range(tables_in_log):
                            h = header_indices[ix]
                            nxt = header_indices[ix + 1]
                            table_start = time.perf_counter()

                            mo_name_from_line = extract_mo_from_subnetwork_line(lines[h])
                            desired_sheet = mo_name_from_line if mo_name_from_line else os.path.splitext(base_filename)[0]

                            df = parse_table_slice_from_subnetwork(lines, h, nxt)
                            note = "Slice parsed"
                            if encoding_used:
                                note += f" | encoding={encoding_used}"
                            df, note = cap_rows(df, note)

                            table_elapsed = time.perf_counter() - table_start
                            if table_elapsed >= float(slow_file_seconds_threshold):
                                any_slow = True

                            mo_names.append(desired_sheet)

                            idx_in_file = per_file_table_idx.get(base_filename, 0)
                            per_file_table_idx[base_filename] = idx_in_file + 1

                            table_entries.append({"df": df, "sheet_candidate": desired_sheet, "log_file": base_filename, "tables_in_log": tables_in_log, "note": note or "", "idx_in_file": idx_in_file})

                        file_elapsed = time.perf_counter() - file_start

                        if show_phase_timings:
                            tag = "[SLOW]" if any_slow or (file_elapsed >= float(slow_file_seconds_threshold)) else ""
                            unique_mo_names: List[str] = []
                            seen = set()
                            for n in mo_names:
                                if n and n not in seen:
                                    seen.add(n)
                                    unique_mo_names.append(n)

                            mo_name_for_log = unique_mo_names[0] if unique_mo_names else "MO NOT FOUND"
                            _log_info(f"PHASE 1: Parse all log/txt files - MO parse {file_counter:>3}: '{mo_name_for_log}' (File: '{base_filename}' ({tables_in_log} tables)) --> took {file_elapsed:.3f}s {tag}")

            # =====================================================================
            #                PHASE 2: Determine final sorting order
            # =====================================================================
            with log_phase_timer("PHASE 2: Determine final sorting order", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                def entry_sort_key(entry: Dict[str, object]) -> Tuple[int, int, int]:
                    """
                    Final sorting key for Excel sheets:
                      - If TABLES_ORDER exists → sort by table order first, then by file (natural), then by table index
                      - Otherwise → sort only by file (natural) and table index
                    """
                    if tables_order:
                        mo = str(entry["sheet_candidate"]).strip()
                        mo_pos = mo_rank.get(mo, len(mo_rank) + 1)
                        return (mo_pos, file_rank.get(entry["log_file"], 10 ** 9), int(entry["idx_in_file"]))
                    return (file_rank.get(entry["log_file"], 10 ** 9), int(entry["idx_in_file"]), 0)

                table_entries.sort(key=entry_sort_key)

            # =====================================================================
            #           PHASE 2.1: Merge repeated MO tables into one sheet
            # =====================================================================
            with log_phase_timer("PHASE 2.1: Merge repeated MO tables", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                merged_entries: List[Dict[str, object]] = []
                by_mo: Dict[str, Dict[str, object]] = {}

                for entry in table_entries:
                    mo_name = str(entry.get("sheet_candidate", "")).strip()
                    if not mo_name:
                        merged_entries.append(entry)
                        continue

                    if mo_name not in by_mo:
                        entry["log_files"] = [str(entry.get("log_file", "")).strip()]
                        entry["notes_list"] = [str(entry.get("note", "")).strip()] if str(entry.get("note", "")).strip() else []
                        by_mo[mo_name] = entry
                        merged_entries.append(entry)
                        continue

                    base = by_mo[mo_name]
                    base_df = base.get("df", pd.DataFrame())
                    new_df = entry.get("df", pd.DataFrame())

                    try:
                        base["df"] = pd.concat([base_df, new_df], ignore_index=True, sort=False)
                    except Exception:
                        base["df"] = base_df

                    lf = str(entry.get("log_file", "")).strip()
                    if lf:
                        base.setdefault("log_files", [])
                        if lf not in base["log_files"]:
                            base["log_files"].append(lf)

                    note = str(entry.get("note", "")).strip()
                    if note:
                        base.setdefault("notes_list", [])
                        if note not in base["notes_list"]:
                            base["notes_list"].append(note)

                    base["note"] = " | ".join([n for n in base.get("notes_list", []) if n])
                    base["tables_in_log"] = max(int(base.get("tables_in_log", 0)), int(entry.get("tables_in_log", 0)))

                table_entries = merged_entries


            # =====================================================================
            #                PHASE 3: Assign unique sheet names
            # =====================================================================
            with log_phase_timer("PHASE 3: Assign unique sheet names", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                used_sheet_names: set = {"Summary", "SummaryAudit"}

                for entry in table_entries:
                    base_name = sanitize_sheet_name(str(entry["sheet_candidate"]))
                    final_sheet = unique_sheet_name(base_name, used_sheet_names)
                    used_sheet_names.add(final_sheet)
                    entry["final_sheet"] = final_sheet

                candidate_to_final_sheet: Dict[str, str] = {
                    str(e.get("sheet_candidate", "")).strip(): str(e.get("final_sheet", "")).strip()
                    for e in table_entries
                    if str(e.get("sheet_candidate", "")).strip() and str(e.get("final_sheet", "")).strip()
                }

            # =====================================================================
            #                PHASE 4: SummaryAudit
            # =====================================================================
            _log_info(f"PHASE 4: SummaryAudit (this phase can take some time)...")
            # =====================================================================
            #                PHASE 4.1: Build SummaryAudit sheet
            # =====================================================================
            with log_phase_timer("PHASE 4.1: Build Summary rows", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                summary_rows: List[Dict[str, object]] = []
                for entry in table_entries:
                    note = str(entry.get("note", ""))
                    separator_str, encoding_str = "", ""

                    # Split "Header=..., | encoding=..." into two separate columns
                    if note:
                        parts = [p.strip() for p in note.split("|")]
                        for part in parts:
                            pl = part.lower()
                            if pl.startswith("header=") or "separated" in pl:
                                separator_str = part
                            elif pl.startswith("encoding="):
                                encoding_str = part.replace("encoding=", "")

                    df: pd.DataFrame = entry["df"]

                    logs_list = entry.get("log_files", [entry.get("log_file", "")])
                    logs_list = [str(x).strip() for x in logs_list if str(x).strip()]
                    logs_list = list(dict.fromkeys(logs_list))  # unique, preserve order

                    if source_zip_path:
                        zip_disp = pretty_path(source_zip_path)
                        logpath_disp = ", ".join([f"{zip_disp}/{lf}" for lf in logs_list]) if logs_list else zip_disp
                    else:
                        logpath_disp = pretty_path(input_dir)

                    summary_rows.append(
                        {
                            "File": ", ".join(logs_list),
                            "Sheet": entry["final_sheet"],
                            "Rows": int(len(df)),
                            "Columns": int(df.shape[1]),
                            "Separator": separator_str,
                            "Encoding": encoding_str,
                            "LogFile": ", ".join(logs_list),
                            "LogPath": logpath_disp,
                            "TablesInLog": entry["tables_in_log"],
                        }
                    )

            # =====================================================================
            #        PHASE 4.2: Prepare pivot tables for extra summary sheets
            # =====================================================================
            with log_phase_timer("PHASE 4.2: Prepare pivot tables", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                # Local Helper to add columns LowMidBand/mmWave to Summary NR_CellDU
                def add_lowmid_mmwave_to_nr_celldu(pivot_df: pd.DataFrame) -> pd.DataFrame:
                    """
                    Add LowMidBand and mmWave columns to Summary NR_CellDU pivot.

                    Logic:
                      - Columns whose header (int) is in [2_000_000, 2_300_000] are mmWave SSBs.
                      - Other numeric SSB columns are LowMidBand SSBs.
                      - For each NodeId we count how many cells it has for each band.
                      - Columns are inserted just before 'Total'.
                    """
                    if pivot_df is None or pivot_df.empty:
                        return pivot_df

                    cols = list(pivot_df.columns)
                    base_cols = {"NodeId", "Total", "LowMidBand", "mmWave"}

                    ssb_cols: list[str] = [c for c in cols if c not in base_cols]

                    mmwave_cols: list[str] = []
                    lowmid_cols: list[str] = []
                    for col in ssb_cols:
                        try:
                            ssb_val = int(str(col))
                        except ValueError:
                            # Non-numeric headers are ignored
                            continue
                        if 2_000_000 <= ssb_val <= 2_300_000:
                            mmwave_cols.append(col)
                        else:
                            lowmid_cols.append(col)

                    lowmid_series = pivot_df[lowmid_cols].sum(axis=1).astype(int) if lowmid_cols else pd.Series(0, index=pivot_df.index, dtype=int)
                    mmwave_series = pivot_df[mmwave_cols].sum(axis=1).astype(int) if mmwave_cols else pd.Series(0, index=pivot_df.index, dtype=int)

                    total_idx = pivot_df.columns.get_loc("Total") if "Total" in pivot_df.columns else len(pivot_df.columns)
                    pivot_df.insert(total_idx, "LowMidBand", lowmid_series)
                    pivot_df.insert(total_idx + 1, "mmWave", mmwave_series)
                    return pivot_df

                # Collect dataframes for the specific MOs we need
                mo_collectors: Dict[str, List[pd.DataFrame]] = {
                    "NRFrequency": [],
                    "NRFreqRelation": [],
                    "NRSectorCarrier": [],
                    "NRCellDU": [],
                    "NRCellRelation": [],
                    "GUtranSyncSignalFrequency": [],
                    "GUtranFreqRelation": [],
                    "GUtranCellRelation": [],
                    "FreqPrioNR": [],
                    "EndcDistrProfile": [],
                    "MeContext": [],
                    # Consistency Checks Post Step2
                    "NRCellCU": [],
                    "EUtranFreqRelation": [],
                    "ExternalNRCellCU": [],
                    "ExternalGUtranCell": [],
                    "TermPointToGNodeB": [],
                    "TermPointToGNB": [],
                    "TermPointToENodeB": [],
                    # <<< NEW: Profiles tables collectors >>>
                    "McpcPCellNrFreqRelProfileUeCfg": [],
                    "McpcPCellProfileUeCfg": [],
                    "UlQualMcpcMeasCfg": [],
                    "McpcPSCellProfileUeCfg": [],
                    "McfbCellProfile": [],
                    "McfbCellProfileUeCfg": [],
                    "TrStSaCellProfile": [],
                    "TrStSaCellProfileUeCfg": [],
                    "CaCellProfile": [],
                    "CaCellProfileUeCfg": [],
                    "TrStSaNrFreqRelProfileUeCfg": [],
                    "McpcPCellEUtranFreqRelProfile": [],
                    "McpcPCellEUtranFreqRelProfileUeCfg": [],
                    "UeMCEUtranFreqRelProfile": [],
                    "UeMCEUtranFreqRelProfileUeCfg": [],
                }
                for entry in table_entries:
                    mo_name = str(entry.get("sheet_candidate", "")).strip()
                    if mo_name in mo_collectors:
                        df_mo = entry["df"]
                        if isinstance(df_mo, pd.DataFrame) and not df_mo.empty:
                            mo_collectors[mo_name].append(df_mo)

                # ---- MeContext must be processed first (UNSYNCHRONIZED nodes must be excluded from all audits) ----
                df_mecontext = concat_or_empty(mo_collectors["MeContext"])

                def _find_col_ci(df: pd.DataFrame, names: List[str]) -> str | None:
                    if df is None or df.empty:
                        return None
                    cols_l = {str(c).strip().lower(): c for c in df.columns}
                    for n in names:
                        key = str(n).strip().lower()
                        if key in cols_l:
                            return cols_l[key]
                    return None

                unsync_nodes: set[str] = set()
                me_node_col = _find_col_ci(df_mecontext, ["NodeId"])
                me_sync_col = _find_col_ci(df_mecontext, ["syncStatus"])
                if df_mecontext is not None and not df_mecontext.empty and me_node_col and me_sync_col:
                    try:
                        mask_unsync = df_mecontext[me_sync_col].astype(str).str.upper().eq("UNSYNCHRONIZED")
                        unsync_nodes = set(df_mecontext.loc[mask_unsync, me_node_col].astype(str).unique())
                    except Exception:
                        unsync_nodes = set()

                def _exclude_unsync(df: pd.DataFrame) -> pd.DataFrame:
                    if df is None or df.empty or not unsync_nodes:
                        return df
                    node_col = _find_col_ci(df, ["NodeId"])
                    if not node_col:
                        return df
                    try:
                        return df.loc[~df[node_col].astype(str).isin(unsync_nodes)].copy()
                    except Exception:
                        return df

                if unsync_nodes:
                    # Filter the already-read sheets (so Excel output also excludes UNSYNCHRONIZED nodes, except MeContext)
                    for entry in table_entries:
                        cand = str(entry.get("sheet_candidate", "")).strip()
                        if cand != "MeContext":
                            df_entry = entry.get("df", None)
                            if isinstance(df_entry, pd.DataFrame) and not df_entry.empty:
                                entry["df"] = _exclude_unsync(df_entry)

                    # Filter collectors used by the audits/pivots
                    for mo_name, dfs in mo_collectors.items():
                        if mo_name != "MeContext" and dfs:
                            mo_collectors[mo_name] = [_exclude_unsync(d) for d in dfs if isinstance(d, pd.DataFrame)]

                # ---- Build pivots ----
                df_nr_cell_du = concat_or_empty(mo_collectors["NRCellDU"])

                # NEW: In NRCellDU, ssbFrequency can be 0 while the real SSB is stored in ssbFrequencyAutoSelected.
                # Replace ssbFrequency=0 by ssbFrequencyAutoSelected so ALL downstream checks/pivots use the real SSB.
                try:
                    if df_nr_cell_du is not None and not df_nr_cell_du.empty:
                        freq_col = next((c for c in df_nr_cell_du.columns if str(c).strip().lower() == "ssbfrequency"), None)
                        auto_col = next((c for c in df_nr_cell_du.columns if str(c).strip().lower() == "ssbfrequencyautoselected"), None)

                        if freq_col and auto_col:
                            freq_num = pd.to_numeric(df_nr_cell_du[freq_col], errors="coerce")
                            auto_num = pd.to_numeric(df_nr_cell_du[auto_col], errors="coerce")

                            mask = (freq_num.fillna(0) == 0) & auto_num.notna() & (auto_num != 0)
                            if bool(mask.any()):
                                df_nr_cell_du.loc[mask, freq_col] = auto_num.loc[mask].astype(int)
                except Exception:
                    pass

                pivot_nr_cells_du = safe_pivot_count(df=df_nr_cell_du, index_field="NodeId", columns_field="ssbFrequency", values_field="NRCellDUId", add_margins=True, margins_name="Total")
                pivot_nr_cells_du = apply_frequency_column_filter(pivot_nr_cells_du, freq_filters)
                pivot_nr_cells_du = add_lowmid_mmwave_to_nr_celldu(pivot_nr_cells_du)

                df_nr_sector_carrier = concat_or_empty(mo_collectors["NRSectorCarrier"])
                pivot_nr_sector_carrier = safe_pivot_count(df=df_nr_sector_carrier, index_field="NodeId", columns_field="arfcnDL", values_field="NRSectorCarrierId", add_margins=True, margins_name="Total")
                pivot_nr_sector_carrier = apply_frequency_column_filter(pivot_nr_sector_carrier, freq_filters)

                df_nr_freq = concat_or_empty(mo_collectors["NRFrequency"])
                pivot_nr_freq = safe_pivot_count(df=df_nr_freq, index_field="NodeId", columns_field="arfcnValueNRDl", values_field="NRFrequencyId", add_margins=True, margins_name="Total")
                pivot_nr_freq = apply_frequency_column_filter(pivot_nr_freq, freq_filters)

                df_nr_freq_rel = concat_or_empty(mo_collectors["NRFreqRelation"])
                pivot_nr_freq_rel = safe_pivot_count(df=df_nr_freq_rel, index_field="NodeId", columns_field="NRFreqRelationId", values_field="NRCellCUId", add_margins=True, margins_name="Total")
                pivot_nr_freq_rel = apply_frequency_column_filter(pivot_nr_freq_rel, freq_filters)

                df_gu_sync_signal_freq = concat_or_empty(mo_collectors["GUtranSyncSignalFrequency"])
                pivot_gu_sync_signal_freq = safe_crosstab_count(df=df_gu_sync_signal_freq, index_field="NodeId", columns_field="arfcn", add_margins=True, margins_name="Total")
                pivot_gu_sync_signal_freq = apply_frequency_column_filter(pivot_gu_sync_signal_freq, freq_filters)

                df_gu_freq_rel = concat_or_empty(mo_collectors["GUtranFreqRelation"])
                pivot_gu_freq_rel = safe_crosstab_count(df=df_gu_freq_rel, index_field="NodeId", columns_field="GUtranFreqRelationId", add_margins=True, margins_name="Total")
                pivot_gu_freq_rel = apply_frequency_column_filter(pivot_gu_freq_rel, freq_filters)

                # Extra tables for audit logic
                df_nr_cell_rel = concat_or_empty(mo_collectors["NRCellRelation"])
                df_gu_cell_rel = concat_or_empty(mo_collectors["GUtranCellRelation"])
                df_freq_prio_nr = concat_or_empty(mo_collectors["FreqPrioNR"])
                df_endc_distr_profile = concat_or_empty(mo_collectors["EndcDistrProfile"])
                df_external_nr_cell_cu = concat_or_empty(mo_collectors["ExternalNRCellCU"])
                df_external_gutran_cell = concat_or_empty(mo_collectors["ExternalGUtranCell"])
                df_term_point_to_gnodeb = concat_or_empty(mo_collectors["TermPointToGNodeB"])
                df_term_point_to_gnb = concat_or_empty(mo_collectors["TermPointToGNB"])
                df_term_point_to_enodeb = concat_or_empty(mo_collectors["TermPointToENodeB"])

                # Extra tables for Consistency Checks Post Step2
                df_nr_cell_cu = concat_or_empty(mo_collectors["NRCellCU"])
                df_eutran_freq_rel = concat_or_empty(mo_collectors["EUtranFreqRelation"])

                # <<< NEW: Build profiles tables dict (only used when profiles_audit=True) >>>
                profile_table_names = [
                    "McpcPCellProfileUeCfg",
                    "McpcPCellNrFreqRelProfileUeCfg",
                    "UlQualMcpcMeasCfg",
                    "McpcPSCellProfileUeCfg",
                    "McfbCellProfile",
                    "McfbCellProfileUeCfg",
                    "McpcPCellEUtranFreqRelProfile",
                    "McpcPCellEUtranFreqRelProfileUeCfg",
                    "UeMCEUtranFreqRelProfile",
                    "UeMCEUtranFreqRelProfileUeCfg",
                    "TrStSaCellProfile",
                    "TrStSaCellProfileUeCfg",
                    "CaCellProfile",
                    "CaCellProfileUeCfg",
                    "TrStSaNrFreqRelProfileUeCfg",
                ]
                profiles_tables: Dict[str, pd.DataFrame] = {}
                if profiles_audit:
                    for table_name in profile_table_names:
                        profiles_tables[table_name] = concat_or_empty(mo_collectors.get(table_name, []))

            # =====================================================================
            #                PHASE 4.3: Build SummaryAudit
            # =====================================================================
            with log_phase_timer("PHASE 4.3: Build SummaryAudit", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                _log_info(f"PHASE 4.3: Build SummaryAudit (this phase can take some time)...")
                summary_audit_df, param_mismatch_nr_df, param_mismatch_gu_df = build_summary_audit(
                    df_mecontext=df_mecontext,
                    df_nr_cell_du=df_nr_cell_du,
                    df_nr_freq=df_nr_freq,
                    df_nr_freq_rel=df_nr_freq_rel,
                    df_nr_cell_rel=df_nr_cell_rel,
                    df_freq_prio_nr=df_freq_prio_nr,
                    df_gu_sync_signal_freq=df_gu_sync_signal_freq,
                    df_gu_freq_rel=df_gu_freq_rel,
                    df_gu_cell_rel=df_gu_cell_rel,
                    df_nr_sector_carrier=df_nr_sector_carrier,
                    df_endc_distr_profile=df_endc_distr_profile,
                    df_nr_cell_cu=df_nr_cell_cu,
                    df_eutran_freq_rel=df_eutran_freq_rel,
                    n77_ssb_pre=self.N77_SSB_PRE,
                    n77_ssb_post=self.N77_SSB_POST,
                    n77b_ssb=self.N77B_SSB,
                    allowed_n77_ssb_pre=self.ALLOWED_N77_SSB_PRE,
                    allowed_n77_arfcn_pre=self.ALLOWED_N77_ARFCN_PRE,
                    allowed_n77_ssb_post=self.ALLOWED_N77_SSB_POST,
                    allowed_n77_arfcn_post=self.ALLOWED_N77_ARFCN_POST,
                    df_external_nr_cell_cu=df_external_nr_cell_cu,
                    df_external_gutran_cell=df_external_gutran_cell,
                    df_term_point_to_gnodeb=df_term_point_to_gnodeb,
                    df_term_point_to_gnb=df_term_point_to_gnb,
                    df_term_point_to_enodeb=df_term_point_to_enodeb,
                    module_name=module_name,
                    profiles_tables=profiles_tables if profiles_audit else None,
                    profiles_audit=profiles_audit,
                    frequency_audit=frequency_audit,
                )

                # Cache in-memory outputs for callers that want to avoid re-reading the Excel from disk (e.g., ConsistencyChecks)
                self._last_summary_audit_df = summary_audit_df
                self._last_param_mismatch_nr_df = param_mismatch_nr_df
                self._last_param_mismatch_gu_df = param_mismatch_gu_df

            # ------------------------------------------------------------------
            # Re-inject modified audit tables back into table_entries
            # ------------------------------------------------------------------
            with log_phase_timer("PHASE 4.4: Re-inject modified tables", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                reinject_map = {
                    "NRCellDU": df_nr_cell_du,
                    "NRFrequency": df_nr_freq,
                    "NRFreqRelation": df_nr_freq_rel,
                    "NRCellRelation": df_nr_cell_rel,
                    "FreqPrioNR": df_freq_prio_nr,
                    "GUtranSyncSignalFrequency": df_gu_sync_signal_freq,
                    "GUFreqRelation": df_gu_freq_rel,
                    "GUtranCellRelation": df_gu_cell_rel,
                    "NRSectorCarrier": df_nr_sector_carrier,
                    "EndcDistrProfile": df_endc_distr_profile,
                    "NRCellCU": df_nr_cell_cu,
                    "EUtranFreqRelation": df_eutran_freq_rel,
                    "ExternalNRCellCU": df_external_nr_cell_cu,
                    "ExternalGUtranCell": df_external_gutran_cell,
                    "TermPointToGNodeB": df_term_point_to_gnodeb,
                    "TermPointToGNB": df_term_point_to_gnb,
                    "TermPointToENodeB": df_term_point_to_enodeb,
                }

                # IMPORTANT: Some MOs can appear multiple times across multiple logs/slices.
                # The aggregated dataframe (concat_or_empty) must be written only once, otherwise we duplicate full content in multiple sheets.
                reinjected_once: set[str] = set()

                for entry in table_entries:
                    sheet_name = str(entry.get("sheet_candidate", "")).strip()
                    entry["skip_write"] = False

                    if sheet_name in reinject_map:
                        if sheet_name in reinjected_once:
                            entry["skip_write"] = True
                            continue

                        entry["df"] = reinject_map[sheet_name]
                        reinjected_once.add(sheet_name)

            # =====================================================================
            #                PHASE 5: Write the Excel file
            # =====================================================================
            with log_phase_timer("PHASE 5: Write Excel", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                tmp_dir, tmp_excel_path = _make_temp_xlsx_path(excel_path_long)
                tmp_excel_path_long = to_long_path(tmp_excel_path)

                try:
                    # ------------------------------------------------------------------
                    # PHASE 5.0: Open ExcelWriter (measure open cost)
                    # ------------------------------------------------------------------
                    t_open0 = time.perf_counter()
                    _log_info(f"PHASE 5.0: ExcelWriter OPEN starting → tmp: '{pretty_path(tmp_excel_path)}'")

                    writer = None
                    writer_closed = False
                    try:
                        excel_engine = "openpyxl"
                        if fast_excel_export:
                            try:
                                import xlsxwriter  # noqa: F401
                                excel_engine = "xlsxwriter"
                            except Exception as e:
                                _log_info(f"PHASE 5.0: fast_excel_export requested but xlsxwriter is not available ({e}). Falling back to openpyxl.")
                                excel_engine = "openpyxl"

                        if excel_engine == "xlsxwriter":
                            # NOTE: Do NOT enable constant_memory when using pandas.to_excel().
                            # Pandas writes cells in an order that violates xlsxwriter constant_memory constraints (row-ordered writes only).
                            # If enabled, sheets end up with only the first column written (and sometimes the last row looks "complete").
                            xlsxwriter_options = {"strings_to_urls": False, "strings_to_numbers": False}
                            writer = pd.ExcelWriter(tmp_excel_path_long, engine="xlsxwriter", engine_kwargs={"options": xlsxwriter_options})
                        else:
                            writer = pd.ExcelWriter(tmp_excel_path_long, engine="openpyxl")

                        written_sheet_dfs: dict[str, pd.DataFrame] = {}
                        _log_info(f"PHASE 5.0: Using Excel engine: {excel_engine}")

                        t_open1 = time.perf_counter()
                        _log_info(f"PHASE 5.0: ExcelWriter OPEN done in {format_duration_hms(t_open1 - t_open0)} ({t_open1 - t_open0:.3f}s)")

                        # We emulate "with" to measure close precisely
                        # (so we can time writer.close() and/or writer.__exit__ behavior)

                        # ---- Enrich MeContext table with additional audit columns (slide 3) ----
                        def _count_by_node(df: pd.DataFrame, node_col: str, cond_series: pd.Series) -> pd.Series:
                            if df is None or df.empty:
                                return pd.Series(dtype=int)
                            try:
                                tmp = df.loc[cond_series.fillna(False), [node_col]].copy()
                                tmp[node_col] = tmp[node_col].astype(str)
                                return tmp.groupby(node_col).size().astype(int)
                            except Exception:
                                return pd.Series(dtype=int)

                        def _format_numeric_like(value) -> str:
                            if value is None or (isinstance(value, float) and pd.isna(value)):
                                return ""
                            txt = str(value).strip()
                            if not txt:
                                return ""
                            try:
                                num = float(txt)
                                if num.is_integer():
                                    return str(int(num))
                                return str(num)
                            except Exception:
                                return txt

                        def _aggregate_by_node(
                            df: pd.DataFrame,
                            node_col: str,
                            relation_col: str,
                            relation_token: int,
                            value_builder,
                        ) -> pd.Series:
                            if df is None or df.empty or not node_col or not relation_col:
                                return pd.Series(dtype=object)

                            work = df.copy()
                            work[node_col] = work[node_col].astype(str)
                            mask = work[relation_col].astype(str).str.contains(str(relation_token), na=False)
                            if not mask.any():
                                return pd.Series(dtype=object)

                            subset = work.loc[mask, [node_col]].copy()
                            subset["_v"] = work.loc[mask].apply(value_builder, axis=1)
                            subset["_v"] = subset["_v"].fillna("").astype(str).str.strip()
                            subset = subset[subset["_v"] != ""]
                            if subset.empty:
                                return pd.Series(dtype=object)

                            def _collapse(vals: pd.Series) -> str:
                                ordered_unique = list(dict.fromkeys(vals.tolist()))
                                return ", ".join(ordered_unique)

                            return subset.groupby(node_col)["_v"].apply(_collapse)

                        if df_mecontext is not None and not df_mecontext.empty:
                            me_node_col = _find_col_ci(df_mecontext, ["NodeId"])
                            me_parent_col = _find_col_ci(df_mecontext, ["ParentId"])
                            if me_node_col:
                                df_me_out = df_mecontext.copy()
                                df_me_out[me_node_col] = df_me_out[me_node_col].astype(str)

                                # NRCellDU derived counts
                                nr_node_col = _find_col_ci(df_nr_cell_du, ["NodeId"])
                                nr_ssb_col = _find_col_ci(df_nr_cell_du, ["ssbFrequency"])
                                if df_nr_cell_du is not None and not df_nr_cell_du.empty and nr_node_col and nr_ssb_col:
                                    nr_ssb_num = pd.to_numeric(df_nr_cell_du[nr_ssb_col], errors="coerce")
                                    s_mmwave = _count_by_node(df_nr_cell_du, nr_node_col, (nr_ssb_num >= 2_000_000) & (nr_ssb_num <= 2_300_000))
                                    s_lowmid = _count_by_node(df_nr_cell_du, nr_node_col, (nr_ssb_num >= 646600) & (nr_ssb_num <= 660000))
                                    s_n77_old = _count_by_node(df_nr_cell_du, nr_node_col, nr_ssb_num.eq(self.N77_SSB_PRE))
                                    s_n77_new = _count_by_node(df_nr_cell_du, nr_node_col, nr_ssb_num.eq(self.N77_SSB_POST))

                                    df_me_out["mmWave Cells"] = df_me_out[me_node_col].map(s_mmwave).fillna(0).astype(int)
                                    df_me_out["LowMidBand Cells"] = df_me_out[me_node_col].map(s_lowmid).fillna(0).astype(int)
                                    df_me_out["N77 Cells"] = df_me_out[me_node_col].map(s_lowmid).fillna(0).astype(int)
                                    df_me_out["N77A old SSB cells"] = df_me_out[me_node_col].map(s_n77_old).fillna(0).astype(int)
                                    df_me_out["N77A new SSB cells"] = df_me_out[me_node_col].map(s_n77_new).fillna(0).astype(int)
                                else:
                                    df_me_out["mmWave Cells"] = 0
                                    df_me_out["LowMidBand Cells"] = 0
                                    df_me_out["N77 Cells"] = 0
                                    df_me_out["N77A old SSB cells"] = 0
                                    df_me_out["N77A new SSB cells"] = 0

                                # NRFreqRelation derived counts (substring match on NRFreqRelationId)
                                nrfr_node_col = _find_col_ci(df_nr_freq_rel, ["NodeId"])
                                nrfr_id_col = _find_col_ci(df_nr_freq_rel, ["NRFreqRelationId"])
                                if df_nr_freq_rel is not None and not df_nr_freq_rel.empty and nrfr_node_col and nrfr_id_col:
                                    s_old_rel = _count_by_node(df_nr_freq_rel, nrfr_node_col, df_nr_freq_rel[nrfr_id_col].astype(str).str.contains(str(self.N77_SSB_PRE), na=False))
                                    s_new_rel = _count_by_node(df_nr_freq_rel, nrfr_node_col, df_nr_freq_rel[nrfr_id_col].astype(str).str.contains(str(self.N77_SSB_POST), na=False))
                                    df_me_out["NRFreqRelation to old N77A SSB"] = df_me_out[me_node_col].map(s_old_rel).fillna(0).astype(int)
                                    df_me_out["NRFreqRelation to new N77A SSB"] = df_me_out[me_node_col].map(s_new_rel).fillna(0).astype(int)
                                else:
                                    df_me_out["NRFreqRelation to old N77A SSB"] = 0
                                    df_me_out["NRFreqRelation to new N77A SSB"] = 0

                                # GUtranFreqRelation derived counts (substring match on GUtranFreqRelationId)
                                gufr_node_col = _find_col_ci(df_gu_freq_rel, ["NodeId"])
                                gufr_id_col = _find_col_ci(df_gu_freq_rel, ["GUtranFreqRelationId"])
                                if df_gu_freq_rel is not None and not df_gu_freq_rel.empty and gufr_node_col and gufr_id_col:
                                    s_old_gufr = _count_by_node(df_gu_freq_rel, gufr_node_col, df_gu_freq_rel[gufr_id_col].astype(str).str.contains(str(self.N77_SSB_PRE), na=False))
                                    s_new_gufr = _count_by_node(df_gu_freq_rel, gufr_node_col, df_gu_freq_rel[gufr_id_col].astype(str).str.contains(str(self.N77_SSB_POST), na=False))
                                    df_me_out["GUtranFreqRelation to old N77A SSB"] = df_me_out[me_node_col].map(s_old_gufr).fillna(0).astype(int)
                                    df_me_out["GUtranFreqRelation to new N77A SSB"] = df_me_out[me_node_col].map(s_new_gufr).fillna(0).astype(int)
                                else:
                                    df_me_out["GUtranFreqRelation to old N77A SSB"] = 0
                                    df_me_out["GUtranFreqRelation to new N77A SSB"] = 0

                                # NR/GU relation priority details requested for MeContext tab.
                                nr_prio_col = _find_col_ci(df_nr_freq_rel, ["cellReselectionPriority"])
                                nr_subprio_col = _find_col_ci(df_nr_freq_rel, ["cellReselectionSubPriority"])
                                if df_nr_freq_rel is not None and not df_nr_freq_rel.empty and nrfr_node_col and nrfr_id_col and nr_prio_col:
                                    s_nr_old_cell_resel = _aggregate_by_node(
                                        df_nr_freq_rel,
                                        nrfr_node_col,
                                        nrfr_id_col,
                                        self.N77_SSB_PRE,
                                        lambda row: (
                                            f"{_format_numeric_like(row.get(nr_prio_col))}.{_format_numeric_like(row.get(nr_subprio_col))}"
                                            if nr_subprio_col and _format_numeric_like(row.get(nr_subprio_col)) != ""
                                            else _format_numeric_like(row.get(nr_prio_col))
                                        ),
                                    )
                                    s_nr_new_cell_resel = _aggregate_by_node(
                                        df_nr_freq_rel,
                                        nrfr_node_col,
                                        nrfr_id_col,
                                        self.N77_SSB_POST,
                                        lambda row: (
                                            f"{_format_numeric_like(row.get(nr_prio_col))}.{_format_numeric_like(row.get(nr_subprio_col))}"
                                            if nr_subprio_col and _format_numeric_like(row.get(nr_subprio_col)) != ""
                                            else _format_numeric_like(row.get(nr_prio_col))
                                        ),
                                    )
                                else:
                                    s_nr_old_cell_resel = pd.Series(dtype=object)
                                    s_nr_new_cell_resel = pd.Series(dtype=object)

                                gu_prio_col = _find_col_ci(df_gu_freq_rel, ["cellReselectionPriority"])
                                gu_subprio_col = _find_col_ci(df_gu_freq_rel, ["cellReselectionSubPriority"])
                                gu_endc_col = _find_col_ci(df_gu_freq_rel, ["endcB1MeasPriority"])

                                if df_gu_freq_rel is not None and not df_gu_freq_rel.empty and gufr_node_col and gufr_id_col:
                                    s_gu_old_cell_resel = _aggregate_by_node(
                                        df_gu_freq_rel,
                                        gufr_node_col,
                                        gufr_id_col,
                                        self.N77_SSB_PRE,
                                        lambda row: (
                                            f"{_format_numeric_like(row.get(gu_prio_col))}.{_format_numeric_like(row.get(gu_subprio_col))}"
                                            if gu_prio_col and gu_subprio_col and _format_numeric_like(row.get(gu_subprio_col)) != ""
                                            else _format_numeric_like(row.get(gu_prio_col))
                                        ),
                                    ) if gu_prio_col else pd.Series(dtype=object)

                                    s_gu_new_cell_resel = _aggregate_by_node(
                                        df_gu_freq_rel,
                                        gufr_node_col,
                                        gufr_id_col,
                                        self.N77_SSB_POST,
                                        lambda row: (
                                            f"{_format_numeric_like(row.get(gu_prio_col))}.{_format_numeric_like(row.get(gu_subprio_col))}"
                                            if gu_prio_col and gu_subprio_col and _format_numeric_like(row.get(gu_subprio_col)) != ""
                                            else _format_numeric_like(row.get(gu_prio_col))
                                        ),
                                    ) if gu_prio_col else pd.Series(dtype=object)

                                    s_gu_old_endc = _aggregate_by_node(
                                        df_gu_freq_rel,
                                        gufr_node_col,
                                        gufr_id_col,
                                        self.N77_SSB_PRE,
                                        lambda row: _format_numeric_like(row.get(gu_endc_col)),
                                    ) if gu_endc_col else pd.Series(dtype=object)

                                    s_gu_new_endc = _aggregate_by_node(
                                        df_gu_freq_rel,
                                        gufr_node_col,
                                        gufr_id_col,
                                        self.N77_SSB_POST,
                                        lambda row: _format_numeric_like(row.get(gu_endc_col)),
                                    ) if gu_endc_col else pd.Series(dtype=object)
                                else:
                                    s_gu_old_cell_resel = pd.Series(dtype=object)
                                    s_gu_new_cell_resel = pd.Series(dtype=object)
                                    s_gu_old_endc = pd.Series(dtype=object)
                                    s_gu_new_endc = pd.Series(dtype=object)

                                df_me_out["NRFreqRelation to old N77A SSB cellReselPrio"] = df_me_out[me_node_col].map(s_nr_old_cell_resel).fillna("")
                                df_me_out["NRFreqRelation to new N77A SSB cellReselPrio"] = df_me_out[me_node_col].map(s_nr_new_cell_resel).fillna("")
                                df_me_out["GUtranFreqRelation to old N77A SSB cellReselPrio"] = df_me_out[me_node_col].map(s_gu_old_cell_resel).fillna("")
                                df_me_out["GUtranFreqRelation to new N77A SSB cellReselPrio"] = df_me_out[me_node_col].map(s_gu_new_cell_resel).fillna("")
                                df_me_out["GUtranFreqRelation to old N77A SSB EndcPrio"] = df_me_out[me_node_col].map(s_gu_old_endc).fillna("")
                                df_me_out["GUtranFreqRelation to new N77A SSB EndcPrio"] = df_me_out[me_node_col].map(s_gu_new_endc).fillna("")

                                sync_col = _find_col_ci(df_me_out, ["syncStatus"])

                                def _split_unique_values(cell_value: str) -> set[str]:
                                    txt = str(cell_value or "").strip()
                                    if not txt:
                                        return set()
                                    return {v.strip() for v in txt.split(",") if v.strip()}

                                def _normalized_priority_values(cell_value: str) -> set[str]:
                                    values = _split_unique_values(cell_value)
                                    return values if values else {"-"}

                                def _build_step1(row: pd.Series) -> str:
                                    sync_val = str(row.get(sync_col, "")).strip().upper() if sync_col else ""
                                    if sync_val == "UNSYNCHRONIZED":
                                        return "SkipUnsynch"

                                    old_cells = int(row.get("N77A old SSB cells", 0) or 0)
                                    old_gu_rel = int(row.get("GUtranFreqRelation to old N77A SSB", 0) or 0)
                                    old_nr_rel = int(row.get("NRFreqRelation to old N77A SSB", 0) or 0)
                                    if old_cells == 0 and old_gu_rel == 0 and old_nr_rel == 0:
                                        return "SkipNoRels"

                                    old_gu = int(row.get("GUtranFreqRelation to old N77A SSB", 0) or 0)
                                    new_gu = int(row.get("GUtranFreqRelation to new N77A SSB", 0) or 0)
                                    old_nr = int(row.get("NRFreqRelation to old N77A SSB", 0) or 0)
                                    new_nr = int(row.get("NRFreqRelation to new N77A SSB", 0) or 0)

                                    old_cell_resel = _normalized_priority_values(row.get("NRFreqRelation to old N77A SSB cellReselPrio", ""))
                                    new_cell_resel = _normalized_priority_values(row.get("NRFreqRelation to new N77A SSB cellReselPrio", ""))
                                    old_endc = _split_unique_values(row.get("GUtranFreqRelation to old N77A SSB EndcPrio", ""))
                                    new_endc = _split_unique_values(row.get("GUtranFreqRelation to new N77A SSB EndcPrio", ""))

                                    # if old_gu == new_gu and old_nr == new_nr and old_cell_resel == new_cell_resel and old_endc != new_endc:
                                    #     return "Step1Done"

                                    cell_resel_same_or_empty = (old_cell_resel == new_cell_resel) or (not old_cell_resel) or (not new_cell_resel)
                                    endc_not_same_or_empty = (old_endc != new_endc) or (not old_endc) or (not new_endc)
                                    if old_gu == new_gu and old_nr == new_nr and cell_resel_same_or_empty and endc_not_same_or_empty:
                                        return "Step1Done"
                                    if old_gu > new_gu or old_nr > new_nr:
                                        return "Step1"
                                    return "Step1Review"

                                def _build_step2b(row: pd.Series) -> str:
                                    old_cells = int(row.get("N77A old SSB cells", 0) or 0)
                                    old_gu_rel = int(row.get("GUtranFreqRelation to old N77A SSB", 0) or 0)
                                    new_cells = int(row.get("N77A new SSB cells", 0) or 0)
                                    if old_cells > 0:
                                        return "Step2b"
                                    if old_cells == 0 and old_gu_rel == 0 and new_cells > 0:
                                        return "Step2bDone"
                                    if old_cells == 0 and old_gu_rel == 0 and new_cells == 0:
                                        return "Step2bNA"
                                    return "Step2bReview"

                                def _build_step2ac(row: pd.Series) -> str:
                                    old_cells = int(row.get("N77A old SSB cells", 0) or 0)
                                    old_gu_rel = int(row.get("GUtranFreqRelation to old N77A SSB", 0) or 0)
                                    if old_cells == 0 and old_gu_rel == 0:
                                        return "SkipNoRels"

                                    old_endc = _split_unique_values(row.get("GUtranFreqRelation to old N77A SSB EndcPrio", ""))
                                    new_endc = _split_unique_values(row.get("GUtranFreqRelation to new N77A SSB EndcPrio", ""))
                                    if old_endc == {"2"} and new_endc == {"1"}:
                                        return "Step2ac"
                                    if old_endc == {"1"} and new_endc == {"2"}:
                                        return "Step2cDone"
                                    return "Step2cReview"

                                df_me_out["Step1"] = df_me_out.apply(_build_step1, axis=1)
                                df_me_out["Step2b"] = df_me_out.apply(_build_step2b, axis=1)
                                df_me_out["Step2ac"] = df_me_out.apply(_build_step2ac, axis=1)

                                df_me_out["Next Step"] = (
                                    df_me_out[["Step1", "Step2b", "Step2ac"]]
                                    .astype(str)
                                    .apply(lambda r: " + ".join(v for v in r.tolist() if v), axis=1)
                                )

                                # Ensure MeContext is exported with enriched columns
                                for entry in table_entries:
                                    if str(entry.get("sheet_candidate", "")).strip() == "MeContext":
                                        entry["df"] = df_me_out
                                        break


                        # ------------------------------------------------------------------
                        # PHASE 5.1: Write Summary + SummaryAudit + Param mismatch sheets
                        # ------------------------------------------------------------------
                        with log_phase_timer("PHASE 5.1: Write Summary + SummaryAudit", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                            # Write Summary first
                            pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)
                            written_sheet_dfs["Summary"] = pd.DataFrame(summary_rows)

                            # SummaryAudit with high-level checks
                            summary_audit_df.to_excel(writer, sheet_name="SummaryAudit", index=False)
                            written_sheet_dfs["SummaryAudit"] = summary_audit_df

                            # Apply alternating background colors by Category for SummaryAudit sheet
                            # NOTE: apply_alternating_category_row_fills is openpyxl-only (xlsxwriter Worksheet is not subscriptable)
                            ws_summary_audit = writer.sheets.get("SummaryAudit")
                            if excel_engine == "openpyxl" and ws_summary_audit is not None:
                                apply_alternating_category_row_fills(ws_summary_audit, category_header="Category")


                            # New: separate NR / LTE param mismatching sheets
                            if not param_mismatch_nr_df.empty:
                                param_mismatch_nr_df.to_excel(writer, sheet_name="Summary NR Param Mismatching", index=False)
                                written_sheet_dfs["Summary NR Param Mismatching"] = param_mismatch_nr_df

                            if not param_mismatch_gu_df.empty:
                                param_mismatch_gu_df.to_excel(writer, sheet_name="Summary LTE Param Mismatching", index=False)
                                written_sheet_dfs["Summary LTE Param Mismatching"] = param_mismatch_gu_df

                        # ------------------------------------------------------------------
                        # PHASE 5.2: Write pivot summary sheets
                        # ------------------------------------------------------------------
                        with log_phase_timer("PHASE 5.2: Write pivot summary sheets", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                            # Extra summary sheets
                            pivot_nr_cells_du.to_excel(writer, sheet_name="Summary NR_CellDU", index=False)
                            written_sheet_dfs["Summary NR_CellDU"] = pivot_nr_cells_du
                            pivot_nr_sector_carrier.to_excel(writer, sheet_name="Summary NR_SectorCarrier", index=False)
                            written_sheet_dfs["Summary NR_SectorCarrier"] = pivot_nr_sector_carrier
                            pivot_nr_freq.to_excel(writer, sheet_name="Summary NR_Frequency", index=False)
                            written_sheet_dfs["Summary NR_Frequency"] = pivot_nr_freq
                            pivot_nr_freq_rel.to_excel(writer, sheet_name="Summary NR_FreqRelation", index=False)
                            written_sheet_dfs["Summary NR_FreqRelation"] = pivot_nr_freq_rel
                            pivot_gu_sync_signal_freq.to_excel(writer, sheet_name="Summary GU_SyncSignalFrequency", index=False)
                            written_sheet_dfs["Summary GU_SyncSignalFrequency"] = pivot_gu_sync_signal_freq
                            pivot_gu_freq_rel.to_excel(writer, sheet_name="Summary GU_FreqRelation", index=False)
                            written_sheet_dfs["Summary GU_FreqRelation"] = pivot_gu_freq_rel


                        # ------------------------------------------------------------------
                        # PHASE 5.3: Write parsed MO tables
                        # ------------------------------------------------------------------
                        _log_info("PHASE 5.3: Write parsed MO tables (this phase can take some time)...")
                        with log_phase_timer("PHASE 5.3: Write parsed MO tables", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                            # Then write each table in the final determined order
                            written = 0
                            total_to_write = sum(1 for e in table_entries if not bool(e.get("skip_write", False)))

                            for entry in table_entries:
                                if bool(entry.get("skip_write", False)):
                                    continue

                                written += 1
                                sheet_start = time.perf_counter()
                                entry["df"].to_excel(writer, sheet_name=entry["final_sheet"], index=False)
                                written_sheet_dfs[str(entry["final_sheet"])] = entry["df"]
                                sheet_elapsed = time.perf_counter() - sheet_start

                                if show_phase_timings and sheet_elapsed >= float(slow_sheet_seconds_threshold):
                                    _log_info(
                                        f"PHASE 5.3: Write parsed MO tables - Slow sheet write {written}/{total_to_write} "
                                        f"(>{slow_sheet_seconds_threshold}s): '{entry['final_sheet']}' ({entry.get('log_file', '')}) "
                                        f"took {sheet_elapsed:.3f}s"
                                    )

                        # ------------------------------------------------------------------
                        # PHASE 5.4: Style sheets (tabs, headers, autofit, hyperlinks)
                        # ------------------------------------------------------------------
                        with log_phase_timer("PHASE 5.4: Style sheets (tabs, headers, autofit, hyperlinks)", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                            if excel_engine == "xlsxwriter":
                                style_headers_autofilter_and_autofit_xlsxwriter(writer, sheet_dfs=written_sheet_dfs, freeze_header=True, align="left", max_autofit_rows=fast_excel_autofit_rows, max_col_width=fast_excel_autofit_max_width, enable_a1_hyperlinks=True, hyperlink_sheet="SummaryAudit", category_sheet_map=candidate_to_final_sheet)
                            else:
                                # Color the 'Summary*' tabs in green
                                color_summary_tabs(writer, prefix="Summary", rgb_hex="00B050")

                                # Apply header color + auto-fit to all sheets
                                # Optimization: default autofit only scans the first N rows (handled inside style_headers_autofilter_and_autofit)
                                style_headers_autofilter_and_autofit(writer, freeze_header=True, align="left", enable_a1_hyperlink=True, hyperlink_sheet="SummaryAudit", category_sheet_map=candidate_to_final_sheet)

                            # MeContext additional conditional formatting (based on slide requirements)
                            try:
                                me_sheet_name = candidate_to_final_sheet.get("MeContext", "MeContext")
                                me_df = written_sheet_dfs.get(me_sheet_name)
                                ws_me = writer.sheets.get(me_sheet_name)

                                if me_df is not None and ws_me is not None and not me_df.empty:
                                    me_cols = list(me_df.columns)
                                    me_col_map = {str(c): i + 1 for i, c in enumerate(me_cols)}

                                    def _xl_col(col_name: str) -> str | None:
                                        idx = me_col_map.get(col_name)
                                        return get_column_letter(idx) if idx else None

                                    def _add_fmt_equals_xlsx(col_target: str, col_ref: str, hex_color: str) -> None:
                                        ct = _xl_col(col_target)
                                        cr = _xl_col(col_ref)
                                        if not ct or not cr:
                                            return
                                        ws_me.conditional_format(
                                            1,
                                            me_col_map[col_target] - 1,
                                            len(me_df),
                                            me_col_map[col_target] - 1,
                                            {
                                                "type": "formula",
                                                "criteria": f"=${ct}2=${cr}2",
                                                "format": writer.book.add_format({"bg_color": hex_color}),
                                            },
                                        )

                                    if excel_engine == "xlsxwriter":
                                        fmt_yellow = writer.book.add_format({"bg_color": "#FFF2CC"})
                                        fmt_green = writer.book.add_format({"bg_color": "#C6EFCE"})
                                        fmt_red = writer.book.add_format({"bg_color": "#FFC7CE"})
                                        fmt_gray = writer.book.add_format({"bg_color": "#D9D9D9"})

                                        for col_name in [
                                            "N77 Cells",
                                            "N77A old SSB cells",
                                            "NRFreqRelation to old N77A SSB",
                                            "GUtranFreqRelation to old N77A SSB",
                                        ]:
                                            idx = me_col_map.get(col_name)
                                            if idx:
                                                ws_me.conditional_format(1, idx - 1, len(me_df), idx - 1, {"type": "cell", "criteria": ">", "value": 0, "format": fmt_yellow})

                                        _add_fmt_equals_xlsx("N77A new SSB cells", "N77A old SSB cells", "#C6EFCE")
                                        _add_fmt_equals_xlsx("NRFreqRelation to new N77A SSB", "NRFreqRelation to old N77A SSB", "#C6EFCE")
                                        _add_fmt_equals_xlsx("GUtranFreqRelation to new N77A SSB", "GUtranFreqRelation to old N77A SSB", "#C6EFCE")
                                        _add_fmt_equals_xlsx("NRFreqRelation to new N77A SSB cellReselPrio", "NRFreqRelation to old N77A SSB cellReselPrio", "#C6EFCE")

                                        nr_new = _xl_col("NRFreqRelation to new N77A SSB cellReselPrio")
                                        nr_old = _xl_col("NRFreqRelation to old N77A SSB cellReselPrio")
                                        if nr_new and nr_old:
                                            ws_me.conditional_format(1, me_col_map["NRFreqRelation to new N77A SSB cellReselPrio"] - 1, len(me_df), me_col_map["NRFreqRelation to new N77A SSB cellReselPrio"] - 1, {"type": "formula", "criteria": f"=AND(${nr_new}2<>\"\",${nr_old}2<>\"\",${nr_new}2<>${nr_old}2)", "format": fmt_red})

                                        _add_fmt_equals_xlsx("GUtranFreqRelation to new N77A SSB cellReselPrio", "NRFreqRelation to old N77A SSB cellReselPrio", "#C6EFCE")
                                        gu_new_resel = _xl_col("GUtranFreqRelation to new N77A SSB cellReselPrio")
                                        if gu_new_resel and nr_old:
                                            ws_me.conditional_format(1, me_col_map["GUtranFreqRelation to new N77A SSB cellReselPrio"] - 1, len(me_df), me_col_map["GUtranFreqRelation to new N77A SSB cellReselPrio"] - 1, {"type": "formula", "criteria": f"=AND(${gu_new_resel}2<>\"\",${nr_old}2<>\"\",${gu_new_resel}2<>${nr_old}2)", "format": fmt_red})

                                        gu_old_endc = _xl_col("GUtranFreqRelation to old N77A SSB EndcPrio")
                                        gu_new_endc = _xl_col("GUtranFreqRelation to new N77A SSB EndcPrio")
                                        if gu_old_endc and gu_new_endc:
                                            ws_me.conditional_format(1, me_col_map["GUtranFreqRelation to new N77A SSB EndcPrio"] - 1, len(me_df), me_col_map["GUtranFreqRelation to new N77A SSB EndcPrio"] - 1, {"type": "formula", "criteria": f"=AND(${gu_new_endc}2<>\"\",${gu_old_endc}2<>\"\",${gu_new_endc}2=${gu_old_endc}2)", "format": fmt_red})

                                        step1_col = _xl_col("Step1")
                                        if step1_col:
                                            ws_me.conditional_format(1, me_col_map["Step1"] - 1, len(me_df), me_col_map["Step1"] - 1, {"type": "formula", "criteria": f"=${step1_col}2=\"SkipNoRels\"", "format": fmt_gray})

                                        step2ac_col = _xl_col("Step2ac")
                                        if step2ac_col:
                                            ws_me.conditional_format(1, me_col_map["Step2ac"] - 1, len(me_df), me_col_map["Step2ac"] - 1, {"type": "formula", "criteria": f"=ISNUMBER(SEARCH(\"Review\",${step2ac_col}2))", "format": fmt_red})
                                            ws_me.conditional_format(1, me_col_map["Step2ac"] - 1, len(me_df), me_col_map["Step2ac"] - 1, {"type": "formula", "criteria": f"=${step2ac_col}2=\"SkipNoRels\"", "format": fmt_gray})
                                    else:
                                        from openpyxl.formatting.rule import FormulaRule, CellIsRule

                                        for col_name in [
                                            "N77 Cells",
                                            "N77A old SSB cells",
                                            "NRFreqRelation to old N77A SSB",
                                            "GUtranFreqRelation to old N77A SSB",
                                        ]:
                                            cl = _xl_col(col_name)
                                            if cl:
                                                ws_me.conditional_formatting.add(
                                                    f"{cl}2:{cl}{len(me_df)+1}",
                                                    CellIsRule(operator="greaterThan", formula=["0"], stopIfTrue=False, fill=PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")),
                                                )

                                        def _add_fmt_equals_openpyxl(col_target: str, col_ref: str, rgb: str) -> None:
                                            ct = _xl_col(col_target)
                                            cr = _xl_col(col_ref)
                                            if not ct or not cr:
                                                return
                                            ws_me.conditional_formatting.add(
                                                f"{ct}2:{ct}{len(me_df)+1}",
                                                FormulaRule(formula=[f"=${ct}2=${cr}2"], stopIfTrue=False, fill=PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")),
                                            )

                                        _add_fmt_equals_openpyxl("N77A new SSB cells", "N77A old SSB cells", "C6EFCE")
                                        _add_fmt_equals_openpyxl("NRFreqRelation to new N77A SSB", "NRFreqRelation to old N77A SSB", "C6EFCE")
                                        _add_fmt_equals_openpyxl("GUtranFreqRelation to new N77A SSB", "GUtranFreqRelation to old N77A SSB", "C6EFCE")
                                        _add_fmt_equals_openpyxl("NRFreqRelation to new N77A SSB cellReselPrio", "NRFreqRelation to old N77A SSB cellReselPrio", "C6EFCE")
                                        _add_fmt_equals_openpyxl("GUtranFreqRelation to new N77A SSB cellReselPrio", "NRFreqRelation to old N77A SSB cellReselPrio", "C6EFCE")

                                        nr_new = _xl_col("NRFreqRelation to new N77A SSB cellReselPrio")
                                        nr_old = _xl_col("NRFreqRelation to old N77A SSB cellReselPrio")
                                        if nr_new and nr_old:
                                            ws_me.conditional_formatting.add(
                                                f"{nr_new}2:{nr_new}{len(me_df)+1}",
                                                FormulaRule(formula=[f"=AND(${nr_new}2<>\"\",${nr_old}2<>\"\",${nr_new}2<>${nr_old}2)"], stopIfTrue=False, fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")),
                                            )

                                        gu_new_resel = _xl_col("GUtranFreqRelation to new N77A SSB cellReselPrio")
                                        if gu_new_resel and nr_old:
                                            ws_me.conditional_formatting.add(
                                                f"{gu_new_resel}2:{gu_new_resel}{len(me_df)+1}",
                                                FormulaRule(formula=[f"=AND(${gu_new_resel}2<>\"\",${nr_old}2<>\"\",${gu_new_resel}2<>${nr_old}2)"], stopIfTrue=False, fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")),
                                            )

                                        gu_old_endc = _xl_col("GUtranFreqRelation to old N77A SSB EndcPrio")
                                        gu_new_endc = _xl_col("GUtranFreqRelation to new N77A SSB EndcPrio")
                                        if gu_old_endc and gu_new_endc:
                                            ws_me.conditional_formatting.add(
                                                f"{gu_new_endc}2:{gu_new_endc}{len(me_df)+1}",
                                                FormulaRule(formula=[f"=AND(${gu_new_endc}2<>\"\",${gu_old_endc}2<>\"\",${gu_new_endc}2=${gu_old_endc}2)"], stopIfTrue=False, fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")),
                                            )

                                        step1_col = _xl_col("Step1")
                                        if step1_col:
                                            ws_me.conditional_formatting.add(
                                                f"{step1_col}2:{step1_col}{len(me_df)+1}",
                                                FormulaRule(formula=[f"=${step1_col}2=\"SkipNoRels\""], stopIfTrue=False, fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")),
                                            )

                                        step2ac_col = _xl_col("Step2ac")
                                        if step2ac_col:
                                            ws_me.conditional_formatting.add(
                                                f"{step2ac_col}2:{step2ac_col}{len(me_df)+1}",
                                                FormulaRule(formula=[f"=ISNUMBER(SEARCH(\"Review\",${step2ac_col}2))"], stopIfTrue=False, fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")),
                                            )
                                            ws_me.conditional_formatting.add(
                                                f"{step2ac_col}2:{step2ac_col}{len(me_df)+1}",
                                                FormulaRule(formula=[f"=${step2ac_col}2=\"SkipNoRels\""], stopIfTrue=False, fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")),
                                            )
                            except Exception as ex:
                                _log_warn(f"PHASE 5.4: Could not apply MeContext conditional formatting: {ex}")


                        # ------------------------------------------------------------------
                        # PHASE 5.5: CLOSE / FINALIZE workbook
                        # ------------------------------------------------------------------
                        _log_info("PHASE 5.5: ExcelWriter CLOSE starting (this phase can take some time)...")
                        t_close0 = time.perf_counter()

                        writer.close()  # <-- this is where openpyxl can take minutes
                        writer_closed = True

                        t_close1 = time.perf_counter()

                        _log_info(f"PHASE 5.5: ExcelWriter CLOSE done in {format_duration_hms(t_close1 - t_close0)} ({t_close1 - t_close0:.3f}s)")

                    finally:
                        # Safety close (avoid leaked handles if something fails mid-write)
                        try:
                            if writer is not None and not writer_closed:
                                # If already closed, this is a no-op / may raise; ignore.
                                writer.close()
                        except Exception:
                            pass

                    # ----------------------------------------------------------------------
                    # PHASE 5.6: Inspect temp file size after close (helps diagnose)
                    # ----------------------------------------------------------------------
                    try:
                        tmp_size = os.path.getsize(tmp_excel_path_long)
                        _log_info(f"PHASE 5.6: Temp XLSX size: {tmp_size / (1024 * 1024):.2f} MB")
                    except Exception as ex:
                        _log_warn(f"PHASE 5.6: Could not read temp XLSX size: {ex}")

                    # ----------------------------------------------------------------------
                    # PHASE 5.7: Move Excel into final destination (prefer atomic replace)
                    # ----------------------------------------------------------------------
                    with log_phase_timer("PHASE 5.7: Move Excel into destination", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                        # Move into final destination (prefer atomic replace)
                        _move_into_place(tmp_excel_path_long, excel_path_long)

                finally:
                    shutil.rmtree(tmp_dir, ignore_errors=True)

            t_close2 = time.perf_counter()
            _log_info(f"PHASE 5: Wrote Excel with {len(table_entries)} sheet(s) in {format_duration_hms(t_close2 - t_open0)} ({t_close2 - t_open0:.3f}s)")
            _log_info(f"PHASE 5: Wrote Excel with {len(table_entries)} sheet(s) in: '{pretty_path(excel_path)}'")

            # =====================================================================
            #                PHASE 6: Export Correction Commands (ConfigurationAudit)
            # =====================================================================
            if export_correction_cmd:
                with log_phase_timer("PHASE 6: Export Correction Commands", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                    sheet_dfs_map: dict[str, pd.DataFrame] = {str(e.get("final_sheet", "")).strip(): df for e in table_entries if not bool(e.get("skip_write", False)) and str(e.get("final_sheet", "")).strip() and isinstance((df := e.get("df")), pd.DataFrame)}

                    # Export External / TermPoint commands (use in-memory DataFrames to avoid re-reading XLSX)
                    export_external_and_termpoint_commands(excel_path_long, base_output_dir_long, base_folder_name=correction_cmd_folder_name, sheet_dfs=sheet_dfs_map, export_to_zip=True, module_name=module_name)

                    # Export any other sheet containing a 'Correction_Cmd' column (NRCellRelation, GUtranCellRelation, etc.)
                    export_all_sheets_with_correction_commands(excel_path_long, base_output_dir_long, base_folder_name=correction_cmd_folder_name, sheet_dfs=sheet_dfs_map, export_to_zip=True, module_name=module_name,
                                                               exclude_sheets={"Summary", "SummaryAudit", "Summary Param Mismatch NR", "Summary Param Mismatch GU", "ExternalNRCellCU", "ExternalGUtranCell", "TermPointToGNodeB", "TermPointToGNB", "NRCellRelation", "GUtranCellRelation"})
            else:
                _log_info("PHASE 6: Export Correction Commands skipped (export_correction_cmd=False or is a Pre-Audit).")

            # =====================================================================
            #                PHASE 7: Generate PPT textual summary
            # =====================================================================
            with log_phase_timer("PHASE 7: Generate PPT summary", log_fn=_log_info, show_start=show_phase_starts, show_end=False, show_timing=show_phase_timings, line_prefix="", start_level="INFO", end_level="INFO", timing_level="INFO"):
                try:
                    ppt_path = generate_ppt_summary(summary_audit_df, excel_path, module_name)
                    if ppt_path:
                        _log_info(f"PPT summary generated in: '{pretty_path(ppt_path)}'")
                except Exception as ex:
                    # Never fail the whole module just for PPT creation
                    _log_warn(f"PPT summary generation failed: {ex}")

            overall_elapsed = time.perf_counter() - overall_start
            if show_phase_timings:
                _log_info(f"TOTAL ConfigurationAudit.run took {format_duration_hms(overall_elapsed)} ({overall_elapsed:.3f}s)")

        return excel_path
