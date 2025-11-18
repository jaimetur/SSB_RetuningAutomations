# -*- coding: utf-8 -*-

import os
import re
import unicodedata
from typing import List, Optional, Tuple, Dict, Iterable
from openpyxl.styles import Alignment
import pandas as pd

# ============================ IO / TEXT ============================

ENCODINGS_TRY = ["utf-8-sig", "utf-16", "utf-16-le", "utf-16-be", "cp1252", "utf-8"]


def read_text_with_encoding(path: str) -> Tuple[List[str], Optional[str]]:
    for enc in ENCODINGS_TRY:
        try:
            with open(path, "r", encoding=enc, errors="strict") as f:
                return [ln.rstrip("\n") for ln in f], enc
        except Exception:
            continue
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        return [ln.rstrip("\n") for ln in f], None


def read_text_lines(path: str) -> Optional[List[str]]:
    try:
        lines, _ = read_text_with_encoding(path)
        return lines
    except Exception:
        return None


# ============================ PARSING ============================

SUMMARY_RE = re.compile(r"^\s*\d+\s+instance\(s\)\s*$", re.IGNORECASE)


def find_all_subnetwork_headers(lines: List[str]) -> List[int]:
    return [i for i, ln in enumerate(lines) if ln.strip().startswith("SubNetwork")]


def extract_mo_from_subnetwork_line(line: str) -> Optional[str]:
    if not line:
        return None
    if "," in line:
        last = line.strip().split(",")[-1].strip()
        return last or None
    toks = line.strip().split()
    return toks[-1].strip() if toks else None


def detect_data_separator(probe_lines: Iterable[str]) -> Optional[str]:
    probe = [ln for ln in probe_lines if ln.strip() and not SUMMARY_RE.match(ln)]
    if any("\t" in ln for ln in probe):
        return "\t"
    if any("," in ln for ln in probe):
        return ","
    return None


def split_line(line: str, sep: Optional[str]) -> List[str]:
    if sep is None:
        return re.split(r"\s+", line.strip())
    return line.split(sep)


def parse_table_slice_from_subnetwork(lines: List[str], header_idx: int, end_idx: int) -> pd.DataFrame:
    # 1) Data header: primera no vacía ni summary tras SubNetwork
    data_header_idx = None
    for j in range(header_idx + 1, end_idx):
        ln = lines[j]
        if not ln.strip() or SUMMARY_RE.match(ln):
            continue
        data_header_idx = j
        break
    if data_header_idx is None:
        return pd.DataFrame()

    header_line = lines[data_header_idx].strip()

    # 2) Separador dentro del slice
    probe_end = min(end_idx, data_header_idx + 50)
    sep = detect_data_separator(lines[data_header_idx:probe_end])

    # 3) Header cols
    header_cols = [c.strip() for c in (split_line(header_line, sep))]
    header_cols = make_unique_columns(header_cols)

    # 4) Filas
    rows: List[List[str]] = []
    for j in range(data_header_idx + 1, end_idx):
        ln = lines[j]
        if not ln.strip() or SUMMARY_RE.match(ln):
            continue
        parts = [p.strip() for p in split_line(ln, sep)]
        if len(parts) < len(header_cols):
            parts += [""] * (len(header_cols) - len(parts))
        elif len(parts) > len(header_cols):
            parts = parts[:len(header_cols)]
        rows.append(parts)

    df = pd.DataFrame(rows, columns=header_cols)
    df = df.replace({"nan": "", "NaN": "", "None": "", "none": "", "NULL": "", "null": ""}).dropna(how="all")
    for c in df.columns:
        df[c] = df[c].astype(str).str.strip()
    return df


# ============================ DATAFRAME UTILS ============================

def normalize_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for c in out.columns:
        out[c] = (
            out[c].astype(str).str.strip()
            .replace({"nan": "", "NaN": "", "None": "", "none": "", "NULL": "", "null": ""})
        )
    return out


def select_latest_by_date(df: pd.DataFrame, side_value: str) -> pd.DataFrame:
    subset = df[df["Pre/Post"].str.lower() == side_value.lower()]
    if subset.empty:
        return subset
    if "Date" not in subset.columns or (subset["Date"].astype(str).str.len() == 0).all():
        return subset
    subset = subset.copy()
    subset["__Date_dt"] = pd.to_datetime(subset["Date"], format="%Y-%m-%d", errors="coerce")
    max_date = subset["__Date_dt"].max()
    return subset[subset["__Date_dt"] == max_date].drop(columns="__Date_dt")


def make_index_by_keys(df: pd.DataFrame, keys: List[str]) -> pd.DataFrame:
    dfx = df.copy()
    for c in keys:
        if c not in dfx.columns:
            dfx[c] = ""
    dfx["_join_key"] = dfx[keys].agg("||".join, axis=1)
    dfx = dfx.set_index("_join_key", drop=True)
    if dfx.index.has_duplicates:
        dfx = dfx[~dfx.index.duplicated(keep="last")]
    return dfx


# ============================ FREQ UTILS ============================

def _base_series(s: pd.Series) -> pd.Series:
    return s.astype(str).str.split("-", n=1).str[0].fillna("").str.strip()


def extract_gu_freq_base(s: pd.Series) -> pd.Series:
    base = _base_series(s)
    fallback = s.astype(str).str.extract(r"(\d+)", expand=False)
    return base.where(base != "", fallback).fillna("").astype(str)


def extract_nr_freq_base(s: pd.Series) -> pd.Series:
    got = s.astype(str).str.extract(r"NRFreqRelation\s*=\s*(\d+)", expand=False)
    return got.fillna(_base_series(s)).fillna("").astype(str)


def detect_freq_column(table_name: str, columns: List[str]) -> Optional[str]:
    if table_name == "GUtranCellRelation" and "GUtranFreqRelationId" in columns:
        return "GUtranFreqRelationId"
    for c in columns:
        lc = c.lower()
        if "freqrelation" in lc or "freq" in lc:
            return c
    return None


def detect_key_columns(table_name: str, columns: List[str], freq_col: Optional[str]) -> List[str]:
    preferred = {
        "GUtranCellRelation": ["GUtranCellRelationId", "neighborCellRef"],
        "NRCellRelation": ["NRCellRelationId", "neighborCellRef"],
    }
    for cand in preferred.get(table_name, []):
        if cand in columns:
            return [cand]
    id_like = [c for c in columns if c.lower().endswith("id")]
    if freq_col and freq_col in id_like:
        id_like.remove(freq_col)
    id_like = [c for c in id_like if c not in ("Pre/Post", "Date")]
    if id_like:
        return id_like[:2]
    if "neighborCellRef" in columns:
        return ["neighborCellRef"]
    remaining = [c for c in columns if c not in ("Pre/Post", "Date")]
    return remaining[:1] if remaining else []


def enforce_gu_columns(df: Optional[pd.DataFrame]) -> pd.DataFrame:
    cols_required = ["NodeId", "EUtranCellFDDId", "GUtranFreqRelationId", "GUtranCellRelationId"]
    if not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame(columns=cols_required)
    out = df.copy()
    for c in cols_required:
        if c not in out.columns:
            out[c] = ""
    other = [c for c in out.columns if c not in cols_required]
    return out[cols_required + other]


def enforce_nr_columns(df: Optional[pd.DataFrame]) -> pd.DataFrame:
    cols_required = ["NodeId", "NRCellCUId", "NRCellRelationId"]
    if not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame(columns=cols_required)
    out = df.copy()
    for c in cols_required:
        if c not in out.columns:
            out[c] = ""
    other = [c for c in out.columns if c not in cols_required]
    return out[cols_required + other]


# ============================ EXCEL HELPERS ============================

def color_summary_tabs(writer, prefix: str = "Summary", rgb_hex: str = "00B050") -> None:
    """
    Set tab color for every worksheet whose name starts with `prefix`.
    Works with openpyxl-backed ExcelWriter.
    - rgb_hex: 6-hex RGB (e.g., '00B050' = green).
    """
    try:
        wb = writer.book  # openpyxl Workbook
        for ws in wb.worksheets:
            if ws.title.startswith(prefix):
                # Set tab color (expects hex without '#')
                ws.sheet_properties.tabColor = rgb_hex
    except Exception:
        # Hard-fail safe: never break file writing just for coloring tabs
        pass


def enable_header_filters(writer, freeze_header: bool = True, align_left: bool = True) -> None:
    """
    Enable Excel AutoFilter on every worksheet for the used range.
    Optionally freeze the header row (row 1) so data scrolls under it.
    Optionally align all cell text to the left.
    """
    try:
        wb = writer.book  # openpyxl Workbook
        for ws in wb.worksheets:
            # Skip empty sheets safely
            if ws.max_row < 1 or ws.max_column < 1:
                continue

            # Define used range for the filter, from A1 to last used cell
            top_left = ws.cell(row=1, column=1).coordinate
            bottom_right = ws.cell(row=ws.max_row, column=ws.max_column).coordinate
            ws.auto_filter.ref = f"{top_left}:{bottom_right}"

            # Optionally freeze header row
            if freeze_header and ws.max_row >= 2:
                ws.freeze_panes = "A2"

            # Optionally align all text to the left
            if align_left:
                left_alignment = Alignment(horizontal="left", vertical="top")
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                    for cell in row:
                        if cell.value is not None:
                            cell.alignment = left_alignment

    except Exception:
        # Never fail the export just for filters or alignment
        pass



def sanitize_sheet_name(name: str) -> str:
    name = re.sub(r'[:\\/?*\[\]]', "_", name)
    name = name.strip().strip("'")
    return (name or "Sheet")[:31]


def unique_sheet_name(base: str, used: set) -> str:
    if base not in used:
        return base
    for k in range(1, 1000):
        suffix = f" ({k})"
        cand = (base[: max(0, 31 - len(suffix))] + suffix)
        if cand not in used:
            return cand
    i, cand = 1, base
    while cand in used:
        cand = f"{base[:28]}_{i:02d}"
        i += 1
    return cand


def try_read_text_file_with_encoding(path: str) -> Tuple[List[str], Optional[str]]:
    """
    Robust text reader that tries several encodings and returns (lines, encoding_used).
    If it falls back to 'replace' mode, returns (lines, None) to signal that encoding is uncertain.
    """
    encodings = ["utf-8-sig", "utf-16", "utf-16-le", "utf-16-be", "cp1252", "utf-8"]
    for enc in encodings:
        try:
            with open(path, "r", encoding=enc, errors="strict") as f:
                return [ln.rstrip("\n") for ln in f], enc
        except Exception:
            continue
    # last permissive attempt
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        return [ln.rstrip("\n") for ln in f], None


def try_read_text_file_lines(path: str) -> Optional[List[str]]:
    """
    Same as above but returns only the lines (used by PrePostRelations.loaders).
    """
    try:
        lines, _ = try_read_text_file_with_encoding(path)
        return lines
    except Exception:
        return None

# --- NEW HELPERS FOR NATURAL FILES SORTING ---

def _split_stem_counter(stem: str) -> Tuple[str, Optional[int]]:
    """
    Split a filename stem into (base_stem, numeric_counter) when it ends with '(N)'.
    Example:
      'Data_Collection_MKT_188(10)' -> ('Data_Collection_MKT_188', 10)
      'Data_Collection_MKT_188'     -> ('Data_Collection_MKT_188', None)
    """
    m = re.search(r"\((\d+)\)\s*$", stem)
    if m:
        return stem[:m.start()].rstrip(), int(m.group(1))
    return stem, None


def natural_logfile_key(path: str) -> Tuple[str, int, int]:
    """
    Generate a natural sorting key for log/text files with '(N)' suffixes.
    Sorts files by:
      1) Base stem (without '(N)') in lowercase
      2) Files *without* counter first, then those *with* counter
      3) Numeric counter ascending
    Example desired order:
      file.txt, file(1).txt, file(2).txt, ..., file(10).txt, file(11).txt
    """
    base = os.path.basename(path)
    stem, _ = os.path.splitext(base)
    stem_base, counter = _split_stem_counter(stem)
    has_counter_flag = 1 if counter is not None else 0
    return (stem_base.lower(), has_counter_flag, counter or 0)

# --- HELPERS FOR DATE DETECTION ---

import re
from datetime import datetime
from typing import Optional, Iterable


def extract_date(folder_name: str) -> Optional[str]:
    """
    Try to find a valid date inside 'folder_name' in many human formats and return it as 'YYYYMMDD'.
    The detection is "intelligent": it validates year, month and day using datetime.strptime.

    Priority rules for ambiguities:
      1) Prefer interpretations that yield the current year.
      2) Prefer formats with 4-digit years (%Y) over 2-digit years (%y).
      3) Prefer candidates containing separators over compact blobs.

    Supported examples (non exhaustive):
      - 20250103, 2025-01-03, 2025_1_3, 2025.01.03
      - 01-03-2025, 1_3_25, 01.03.25
      - 03/01/2025 (both D/M/Y and M/D/Y are attempted)
      - Jan-03-2025, 03-Jan-25, January 3 2025, 3 January 25
      - 2025Jan03, 03Jan2025, 3January2025
      - With or without separators: '-', '_', '.', '/', ' '

    Two-digit years are mapped with a century window [1970..2069].
    Returns 'YYYYMMDD' on success, otherwise None.
    """

    # ------------------------- helpers -------------------------
    def try_parse_with_formats(candidate: str, fmts: Iterable[str]) -> list[tuple[datetime, str]]:
        """Return all successful parses as (datetime, format) to allow scoring selection."""
        results: list[tuple[datetime, str]] = []
        for fmt in fmts:
            try:
                dt = datetime.strptime(candidate, fmt)
                if 1900 <= dt.year <= 2100:
                    # Enforce 2-digit year window [1970..2069]
                    if "%y" in fmt and not (1970 <= dt.year <= 2069):
                        continue
                    results.append((dt, fmt))
            except ValueError:
                continue
        return results

    def normalize_output(dt: datetime) -> str:
        """Return canonical YYYY-MM-DD string."""
        return dt.strftime("%Y-%m-%d")

    # -------------------- candidate generation --------------------
    text = folder_name
    current_year = datetime.now().year
    current_year_str = str(current_year)
    current_year_2d = f"{current_year % 100:02d}"

    # Tokenize by non-alnum; keep alnum tokens
    tokens = [t for t in re.split(r"[^A-Za-z0-9]+", text) if t]

    candidates: set[str] = set()

    # 1) Individual tokens (useful for '2025Jan03' or numeric blobs)
    for t in tokens:
        if len(t) >= 4:
            candidates.add(t)
        if t.isdigit():
            candidates.add(t)

    # 2) Windows joined with/without separators
    seps = ["-", "_", ".", "/", " "]
    for win_size in (2, 3, 4):
        for i in range(0, len(tokens) - win_size + 1):
            window = tokens[i:i + win_size]
            joined_no_sep = "".join(window)
            if 4 <= len(joined_no_sep) <= 16:
                candidates.add(joined_no_sep)
            for sep in seps:
                joined = sep.join(window)
                if 4 <= len(joined) <= 20:
                    candidates.add(joined)

    # 3) Regex slices that look date-ish
    dateish_regexes = [
        r"\b\d{8}\b",                               # 20250103 / 01032025
        r"\b\d{6}\b",                               # 250103 / 111125
        r"\b\d{4}[-_/.\s]\d{1,2}[-_/.\s]\d{1,2}\b",
        r"\b\d{1,2}[-_/.\s]\d{1,2}[-_/.\s]\d{2,4}\b",
        r"\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[-_/.\s]?\d{1,2}[-_/.\s]?\d{2,4}\b",
        r"\b\d{1,2}[-_/.\s]?(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[-_/.\s]?\d{2,4}\b",
        r"\b\d{4}[-_/.\s]?(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[-_/.\s]?\d{1,2}\b",
        r"\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\d{1,2}\d{4}\b",  # Jan032025
        r"\b\d{1,2}(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\d{4}\b",  # 03Jan2025
    ]
    for rx in dateish_regexes:
        for m in re.finditer(rx, text, flags=re.IGNORECASE):
            candidates.add(m.group(0))

    if not candidates:
        return None

    # -------------------- formats to attempt --------------------
    # Keep a broad set; scoring will decide which parse wins.
    numeric_YMD = [
        "%Y%m%d", "%Y-%m-%d", "%Y_%m_%d", "%Y.%m.%d", "%Y/%m/%d", "%Y %m %d",
        "%y%m%d", "%y-%m-%d", "%y_%m_%d", "%y.%m.%d", "%y/%m/%d", "%y %m %d",
    ]
    numeric_MDY = [
        "%m%d%Y", "%m-%d-%Y", "%m_%d_%Y", "%m.%d.%Y", "%m/%d/%Y", "%m %d %Y",
        "%m%d%y", "%m-%d-%y", "%m_%d_%y", "%m.%d.%y", "%m/%d/%y", "%m %d %y",
    ]
    numeric_DMY = [
        "%d%m%Y", "%d-%m-%Y", "%d_%m_%Y", "%d.%m.%Y", "%d/%m/%Y", "%d %m %Y",
        "%d%m%y", "%d-%m-%y", "%d_%m_%y", "%d.%m.%y", "%d/%m/%y", "%d %m %y",
    ]
    month_name_variants = [
        # Day Month Year
        "%d-%b-%Y", "%d %b %Y", "%d_%b_%Y", "%d.%b.%Y", "%d/%b/%Y",
        "%d-%B-%Y", "%d %B %Y", "%d_%B_%Y", "%d.%B.%Y", "%d/%B/%Y",
        # Month Day Year
        "%b-%d-%Y", "%b %d %Y", "%b_%d_%Y", "%b.%d.%Y", "%b/%d/%Y",
        "%B-%d-%Y", "%B %d %Y", "%B_%d_%Y", "%B.%d.%Y", "%B/%d/%Y",
        # Year Month Day
        "%Y-%b-%d", "%Y %b %d", "%Y_%b_%d", "%Y.%b.%d", "%Y/%b/%d",
        "%Y-%B-%d", "%Y %B %d", "%Y_%B_%d", "%Y.%B.%d", "%Y/%B/%d",
        # Compact without separators like 03Jan2025 / Jan032025 / 2025Jan03
        "%d%b%Y", "%b%d%Y", "%Y%b%d",
        "%d%B%Y", "%B%d%Y", "%Y%B%d",
        # Two-digit year with names
        "%d-%b-%y", "%d %b %y", "%b %d %y", "%y %b %d", "%b%d%y", "%d%b%y",
        "%d-%B-%y", "%d %B %y", "%B %d %y", "%y %B %d", "%B%d%y", "%d%B%y",
    ]
    all_formats: list[str] = numeric_YMD + numeric_MDY + numeric_DMY + month_name_variants

    # -------------------- parsing with scoring --------------------
    # We will evaluate all successful parses per candidate and select the best by a score tuple.
    # Lower score is better.
    def score_parse(candidate: str, dt: datetime, fmt: str) -> tuple[int, int, int, int]:
        """Build a priority score: (year_match, four_digit_year, has_separators, candidate_len)."""
        year_match = 0 if dt.year == current_year else 1
        four_digit = 0 if "%Y" in fmt else 1
        has_seps = 0 if any(sep in candidate for sep in seps) else 1
        # Slight bias towards shorter candidates to avoid over-greedy matches
        return (year_match, four_digit, has_seps, len(candidate))

    best_dt: Optional[datetime] = None
    best_score: Optional[tuple[int, int, int, int]] = None

    # Try candidates that contain the current year first (fast path).
    # This improves cases like "...2025-11-11..." where we clearly want 2025.
    prioritized_candidates = sorted(
        candidates,
        key=lambda c: (current_year_str not in c, len(c))
    )

    for candidate in prioritized_candidates:
        parses = try_parse_with_formats(candidate, all_formats)
        if not parses:
            continue

        # Among all parses for this candidate, pick by score
        for dt, fmt in parses:
            # If the candidate is a pure 6-digit blob, prioritize interpretations that place %y at the END
            # (i.e., mmddyy or ddmmyy) because users often encode '...yy' as the last two digits.
            if candidate.isdigit() and len(candidate) == 6:
                if "%y" in fmt and fmt.endswith("%y"):
                    # small bonus: improve score artificially
                    sc = score_parse(candidate, dt, fmt)
                    sc = (sc[0], sc[1], sc[2], max(sc[3] - 1, 0))
                else:
                    sc = score_parse(candidate, dt, fmt)
            else:
                sc = score_parse(candidate, dt, fmt)

            # Pick the best-scoring parse globally
            if best_score is None or sc < best_score:
                best_score = sc
                best_dt = dt

            # Early exit if perfect score (current year + 4-digit + separators)
            if sc[:3] == (0, 0, 0):
                return normalize_output(dt)

    return normalize_output(best_dt) if best_dt else None

# --- NEW HELPERS FROM CONFIGURATION AUDIT MODULE ---
def resolve_column_case_insensitive(
    df: pd.DataFrame,
    candidates: List[str],
) -> Optional[str]:
    """
    Resolve a column name by trying several candidates, case-insensitive
    and ignoring underscores/spaces.
    """
    if df is None or df.empty:
        return None

    def _canon(s: str) -> str:
        return re.sub(r"[\s_]+", "", str(s).strip().lower())

    cols = list(df.columns)
    canon_map = {_canon(c): c for c in cols}
    for cand in candidates:
        key = _canon(cand)
        if key in canon_map:
            return canon_map[key]
    # Fallback: startswith-based match
    for cand in candidates:
        key = _canon(cand)
        for c in cols:
            if _canon(c).startswith(key):
                return c
    return None


def parse_int_frequency(value: object) -> Optional[int]:
    """
    Try to parse a frequency/ARFCN value as integer from the leading numeric part
    of the string (before any non-digit chars like '-' or spaces).

    Examples:
      - '653952-30-20-0-1' -> 653952
      - '648672 some text' -> 648672
      - '  647328'         -> 647328
    """
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None

    m = re.match(r"^(\d+)", s)
    if not m:
        return None

    try:
        return int(m.group(1))
    except Exception:
        return None


def is_n77_from_string(value: object) -> bool:
    """
    Determine if a cell can be considered N77 based on ARFCN/SSB string.

    Here we approximate N77 as frequencies whose textual representation starts with '6'.
    """
    if value is None:
        return False
    s = str(value).strip()
    return bool(s) and s[0] == "6"


def concat_or_empty(dfs: List[pd.DataFrame]) -> pd.DataFrame:
    """
    Return a single concatenated DataFrame or an empty one if none;
    align on common columns when required.
    """
    if not dfs:
        return pd.DataFrame()
    try:
        return pd.concat(dfs, ignore_index=True)
    except Exception:
        common_cols = set.intersection(*(set(d.columns) for d in dfs)) if dfs else set()
        if not common_cols:
            return pd.DataFrame()
        dfs_aligned = [d[list(common_cols)].copy() for d in dfs]
        return pd.concat(dfs_aligned, ignore_index=True)


def safe_pivot_count(
    df: pd.DataFrame,
    index_field: str,
    columns_field: str,
    values_field: str,
    add_margins: bool = True,
    margins_name: str = "Total",
) -> pd.DataFrame:
    """
    Robust pivot builder that prevents 'Grouper for ... not 1-dimensional' errors.
    """
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame({"Info": ["Table not found or empty"]})

    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [
            "_".join([str(c).strip() for c in tup if str(c).strip()])
            for tup in df.columns
        ]
    if isinstance(df.index, pd.MultiIndex):
        df = df.reset_index()

    work = df.reset_index(drop=True).copy()
    work.columns = pd.Index([str(c).strip() for c in work.columns])

    # Remove duplicate columns (case-insensitive)
    seen_lower = set()
    unique_cols = []
    for c in work.columns:
        cl = c.lower()
        if cl in seen_lower:
            continue
        seen_lower.add(cl)
        unique_cols.append(c)
    work = work[unique_cols]

    def _resolve(name: str) -> Optional[str]:
        nl = name.lower()
        for c in work.columns:
            if c.lower() == nl or c.lower().startswith(nl + "_"):
                return c
        return None

    idx_col = _resolve(index_field)
    col_col = _resolve(columns_field)
    val_col = _resolve(values_field)

    if not all([idx_col, col_col, val_col]):
        missing = [
            n
            for n, v in [
                (index_field, idx_col),
                (columns_field, col_col),
                (values_field, val_col),
            ]
            if v is None
        ]
        return pd.DataFrame(
            {
                "Info": [f"Required columns missing: {', '.join(missing)}"],
                "PresentColumns": [", ".join(work.columns.tolist())],
            }
        )

    for col in {idx_col, col_col, val_col}:
        work[col] = work[col].astype(str).str.strip()

    try:
        piv = pd.pivot_table(
            work,
            index=idx_col,
            columns=col_col,
            values=val_col,
            aggfunc="count",
            fill_value=0,
            margins=add_margins,
            margins_name=margins_name,
        ).reset_index()

        if isinstance(piv.columns, pd.MultiIndex):
            piv.columns = [
                " ".join([str(x) for x in tup if str(x)]).strip()
                for tup in piv.columns
            ]

        return piv

    except Exception as ex:
        return pd.DataFrame(
            {
                "Error": [f"Pivot build failed: {ex}"],
                "PresentColumns": [", ".join(work.columns.tolist())],
            }
        )


def safe_crosstab_count(
    df: pd.DataFrame,
    index_field: str,
    columns_field: str,
    add_margins: bool = True,
    margins_name: str = "Total",
) -> pd.DataFrame:
    """
    Build a frequency table with pd.crosstab (no 'values' column required).
    """
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame({"Info": ["Table not found or empty"]})

    work = df.copy()
    if isinstance(work.columns, pd.MultiIndex):
        work.columns = [
            "_".join([str(c) for c in tup if str(c)]).strip()
            for tup in work.columns
        ]
    if isinstance(work.index, pd.MultiIndex):
        work = work.reset_index()
    work = work.reset_index(drop=True)

    def _norm_header(s: str) -> str:
        s = "" if s is None else str(s)
        s = unicodedata.normalize("NFKC", s).replace("\ufeff", "").replace("\u200b", "").replace("\xa0", " ")
        s = re.sub(r"\s+", " ", s.strip())
        return s

    work.columns = pd.Index([_norm_header(c) for c in work.columns])

    def _canon(s: str) -> str:
        s = s.lower().replace(" ", "").replace("_", "").replace("-", "")
        return s

    seen = set()
    keep = []
    for c in work.columns:
        k = _canon(c)
        if k in seen:
            continue
        seen.add(k)
        keep.append(c)
    work = work[keep]

    def _resolve(name: str) -> Optional[str]:
        target = _canon(_norm_header(name))
        for c in work.columns:
            if _canon(c) == target:
                return c
        for c in work.columns:
            if _canon(c).startswith(target):
                return c
        return None

    idx_col = _resolve(index_field)
    col_col = _resolve(columns_field)
    if not idx_col or not col_col:
        missing = [
            n
            for n, v in [(index_field, idx_col), (columns_field, col_col)]
            if v is None
        ]
        return pd.DataFrame(
            {
                "Info": [f"Required columns missing: {', '.join(missing)}"],
                "PresentColumns": [", ".join(work.columns.tolist())],
            }
        )

    work[idx_col] = work[idx_col].astype(str).map(_norm_header)
    work[col_col] = work[col_col].astype(str).map(_norm_header)

    try:
        ct = pd.crosstab(
            index=work[idx_col],
            columns=work[col_col],
            dropna=False,
        ).reset_index()

        if add_margins:
            ct["Total"] = ct.drop(columns=[idx_col]).sum(axis=1)
            total_row = {idx_col: "Total"}
            for c in ct.columns:
                if c != idx_col:
                    total_row[c] = ct[c].sum()
            ct = pd.concat([ct, pd.DataFrame([total_row])], ignore_index=True)

        return ct
    except Exception as ex:
        return pd.DataFrame(
            {
                "Error": [f"Crosstab build failed: {ex}"],
                "PresentColumns": [", ".join(work.columns.tolist())],
            }
        )


def apply_frequency_column_filter(piv: pd.DataFrame, filters: List[str]) -> pd.DataFrame:
    """
    Keep only the first (index) column, 'Total' (if present), and columns whose
    header contains any of the provided substrings (case-insensitive).
    """
    if not isinstance(piv, pd.DataFrame) or piv.empty or not filters:
        return piv

    cols = [str(c) for c in piv.columns.tolist()]
    keep: List[str] = []

    if cols:
        keep.append(cols[0])

    fl = [f.lower() for f in filters if f]
    for c in cols[1:]:
        lc = c.lower()
        if c == "Total" or lc == "total":
            keep.append(c)
            continue
        if any(f in lc for f in fl):
            keep.append(c)

    if len(keep) <= 1 and "Total" in cols and "Total" not in keep:
        keep.append("Total")

    try:
        return piv.loc[:, keep]
    except Exception:
        return piv


def find_log_files(folder: str) -> List[str]:
    """
    Return a sorted list of *.log / *.logs / *.txt files found in 'folder'.
    """
    files: List[str] = []
    for name in os.listdir(folder):
        lower = name.lower()
        if lower.endswith((".log", ".logs", ".txt")):
            p = os.path.join(folder, name)
            if os.path.isfile(p):
                files.append(p)
    files.sort()
    return files


def read_text_file(path: str) -> Tuple[List[str], Optional[str]]:
    """
    Thin wrapper around read_text_with_encoding to keep current behavior.
    Returns (lines, encoding_used).
    """
    return read_text_with_encoding(path)


def fallback_header_index(
    valid_lines: List[str],
    all_lines: List[str],
    summary_re,
) -> Optional[int]:
    """
    Best-effort detection of a tabular header line when no explicit 'SubNetwork' header is found.
    """
    any_tab = any("\t" in ln for ln in valid_lines)
    sep: Optional[str] = "\t" if any_tab else ("," if any("," in ln for ln in valid_lines) else None)

    for i, ln in enumerate(all_lines):
        if not ln.strip() or summary_re.match(ln):
            continue
        if sep == "\t" and "\t" in ln:
            return i
        if sep == "," and "," in ln:
            return i
        if sep is None:
            # First non-empty, non-summary line becomes header
            return i
    return None


def parse_log_lines(
    lines: List[str],
    summary_re,
    forced_header_idx: Optional[int] = None,
) -> Tuple[pd.DataFrame, str]:
    """
    Parse a list of lines into a DataFrame.

    - Tries to infer separator (tab / comma / whitespace).
    - Skips lines matching SUMMARY_RE.
    - Normalizes columns and trims rows beyond Excel's maximum if needed.
    """
    valid = [ln for ln in lines if ln.strip() and not summary_re.match(ln)]
    header_idx = forced_header_idx
    if header_idx is None:
        header_idx = fallback_header_index(valid, lines, summary_re)
    if header_idx is None:
        return pd.DataFrame(), "No header detected"

    header_line = lines[header_idx].strip()
    any_tab = any("\t" in ln for ln in valid)
    data_sep: Optional[str] = "\t" if any_tab else ("," if any("," in ln for ln in valid) else None)

    if header_line.startswith("SubNetwork"):
        header_cols = [c.strip() for c in header_line.split(",")]
    else:
        header_cols = [
            c.strip()
            for c in (
                header_line.split(data_sep)
                if data_sep
                else re.split(r"\s+", header_line.strip())
            )
        ]
    header_cols = make_unique_columns(header_cols)

    rows: List[List[str]] = []
    for ln in lines[header_idx + 1 :]:
        if not ln.strip() or summary_re.match(ln):
            continue
        parts = [
            p.strip()
            for p in (
                ln.split(data_sep)
                if data_sep
                else re.split(r"\s+", ln.strip())
            )
        ]
        if len(parts) < len(header_cols):
            parts += [""] * (len(header_cols) - len(parts))
        elif len(parts) > len(header_cols):
            parts = parts[: len(header_cols)]
        rows.append(parts)

    df = pd.DataFrame(rows, columns=header_cols)
    df = df.replace({"nan": "", "NaN": "", "None": "", "NULL": ""}).dropna(how="all")
    for c in df.columns:
        df[c] = df[c].astype(str).str.strip()

    note = (
        "Header=SubNetwork-comma"
        if header_line.startswith("SubNetwork")
        else (
            "Tab-separated"
            if data_sep == "\t"
            else ("Comma-separated" if data_sep == "," else "Whitespace-separated")
        )
    )
    return df, note


def find_subnetwork_header_index(lines: List[str], summary_re=None) -> Optional[int]:
    """
    Find the index of the first line starting with 'SubNetwork'.
    """
    for i, ln in enumerate(lines):
        if ln.strip().startswith("SubNetwork"):
            return i
    return None


def extract_mo_name_from_previous_line(
    lines: List[str],
    header_idx: Optional[int],
) -> Optional[str]:
    """
    Try to infer the MO name from the line immediately before the header.
    """
    if header_idx is None or header_idx == 0:
        return None
    prev = lines[header_idx - 1].strip()
    if not prev:
        return None
    if "," in prev:
        last = prev.split(",")[-1].strip()
        return last or None
    toks = prev.split()
    return toks[-1].strip() if toks else None


def cap_rows(
    df: pd.DataFrame,
    note: str,
    max_rows_excel: int = 1_048_576,
) -> Tuple[pd.DataFrame, str]:
    """
    Cap the number of rows to Excel's maximum and append a note if trimming occurs.
    """
    if len(df) > max_rows_excel:
        df = df.iloc[:max_rows_excel, :].copy()
        note = (note + " | " if note else "") + f"Trimmed to {max_rows_excel} rows"
    return df, note


def make_unique_columns(cols: List[str]) -> List[str]:
    """
    Ensure column names are unique by appending a numeric suffix when needed.
    """
    seen: Dict[str, int] = {}
    unique: List[str] = []
    for c in cols:
        base = c or "Col"
        if base not in seen:
            seen[base] = 0
            unique.append(base)
        else:
            seen[base] += 1
            unique.append(f"{base}_{seen[base]}")
    return unique

# def make_unique_columns(cols: List[str]) -> List[str]:
#     seen: Dict[str, int] = {}
#     out = []
#     for c in cols:
#         if c not in seen:
#             seen[c] = 0
#             out.append(c)
#         else:
#             seen[c] += 1
#             out.append(f"{c}.{seen[c]}")
#     return out
