# -*- coding: utf-8 -*-

import os
import re
from typing import Dict, Optional, List

import pandas as pd

from src.modules.Common.correction_commands_builder import build_correction_command_gu_new_relations, build_correction_command_gu_missing_relations, build_correction_command_gu_discrepancies, build_correction_command_nr_new_relations, build_correction_command_nr_missing_relations, build_correction_command_nr_discrepancies
from src.utils.utils_dataframe import select_latest_by_date, normalize_df, make_index_by_keys
from src.utils.utils_datetime import extract_date
from src.utils.utils_excel import color_summary_tabs, style_headers_autofilter_and_autofit, apply_alternating_category_row_fills, style_headers_autofilter_and_autofit_xlsxwriter
from src.utils.utils_frequency import detect_freq_column, detect_key_columns, extract_gu_freq_base, extract_nr_freq_base, enforce_gu_columns, enforce_nr_columns
from src.utils.utils_io import read_text_lines, to_long_path, pretty_path
from src.utils.utils_parsing import find_all_subnetwork_headers, extract_mo_from_subnetwork_line, parse_table_slice_from_subnetwork
from src.modules.Common.common_functions import load_nodes_names_and_id_from_summary_audit
from src.modules.Common.correction_commands_exporter import export_relations_commands


class ConsistencyChecks:
    """
    Loads and compares GU/NR relation tables before (Pre) and after (Post) a refarming process.
    (Se mantiene la funcionalidad exacta.)
    """
    PRE_TOKENS = ("pre",)   # mantenemos la semántica previa (antes: ("pre", "step0"))
    POST_TOKENS = ("post",)
    DATE_RE = re.compile(r"(?P<date>(19|20)\d{6})")  # yyyymmdd
    SUMMARY_RE = re.compile(r"^\s*\d+\s+instance\(s\)\s*$", re.IGNORECASE)


    # ------------------------------------------------------------------
    #  CONSTRUCTOR
    # ------------------------------------------------------------------
    def __init__(
        self,
        n77_ssb_pre: Optional[str] = None,
        n77_ssb_post: Optional[str] = None,
        freq_filter_list: Optional[List[str]] = None,
    ) -> None:
        # NEW: store N77 SSB frequencies for Pre and Post
        self.n77_ssb_pre: Optional[str] = n77_ssb_pre
        self.n77_ssb_post: Optional[str] = n77_ssb_post

        # NEW: optional frequency filter list (strings)
        self.freq_filter_list: List[str] = [
            str(f).strip() for f in (freq_filter_list or []) if str(f).strip()
        ]

        self.tables: Dict[str, pd.DataFrame] = {}

        # NEW: flags to signal whether at least one Pre/Post folder was found
        self.pre_folder_found: bool = False
        self.post_folder_found: bool = False

        # NEW: keep only per-table/per-side source file paths to be used exclusively in Summary (do not store them in DataFrames)
        self._source_paths: Dict[str, Dict[str, List[tuple]]] = {
            "GUtranCellRelation": {"Pre": [], "Post": []},
            "NRCellRelation": {"Pre": [], "Post": []},
        }

        # NEW: keep paths to PRE/POST ConfigurationAudit Excel files
        self.audit_pre_excel: Optional[str] = None
        self.audit_post_excel: Optional[str] = None

        # NEW: keep SummaryAudit DataFrames in memory to avoid re-loading Excel multiple times
        self.audit_pre_summary_audit_df: Optional[pd.DataFrame] = None
        self.audit_post_summary_audit_df: Optional[pd.DataFrame] = None

    # ------------------------------------------------------------------
    #  SHARED SMALL HELPERS (para reducir líneas en funciones repetidas)
    # ------------------------------------------------------------------
    # --------- folder helpers ---------
    @staticmethod
    def _detect_prepost(folder_name: str) -> Optional[str]:
        name = folder_name.lower()
        if any(tok in name for tok in ConsistencyChecks.PRE_TOKENS):
            return "Pre"
        if any(tok in name for tok in ConsistencyChecks.POST_TOKENS):
            return "Post"
        return None

    @staticmethod
    def _insert_front_columns(df: pd.DataFrame, prepost: str, date_str: Optional[str]) -> pd.DataFrame:
        df = df.copy()
        df.insert(0, "Pre/Post", prepost)
        df.insert(1, "Date", date_str if date_str else "")
        return df

    @staticmethod
    def _table_key_name(table_base: str) -> str:
        return table_base.strip()

    def _filter_rows_by_freq_list(self, df: Optional[pd.DataFrame]) -> Optional[pd.DataFrame]:
        """
        Filter rows keeping only those where Freq_Pre or Freq_Post contains
        any of the frequencies in self.freq_filter_list.

        If the filter list is empty, the input DataFrame is returned unchanged.
        """
        if df is None or df.empty:
            return df
        if not self.freq_filter_list:
            return df

        df = df.copy()
        if "Freq_Pre" not in df.columns:
            df["Freq_Pre"] = ""
        if "Freq_Post" not in df.columns:
            df["Freq_Post"] = ""

        pattern = "|".join(re.escape(f) for f in self.freq_filter_list if f)
        if not pattern:
            return df

        combined = df["Freq_Pre"].astype(str) + " " + df["Freq_Post"].astype(str)
        mask = combined.str.contains(pattern, regex=True, na=False)
        return df[mask]

    # ----------------------------- LOADING ----------------------------- #
    def _collect_from_dir(self, dir_path: str, prepost: str, collected: Dict[str, List[pd.DataFrame]]) -> None:
        """
        Small helper used in both legacy and dual-input modes to avoid duplication.
        """
        # Try to extract date from the last 3 folder levels (current, parent, grandparent) safely
        dir_basename = os.path.basename(dir_path)
        parent = os.path.dirname(dir_path)
        grandparent = os.path.dirname(parent) if parent else ""
        parent_basename = os.path.basename(parent) if parent else ""
        grandparent_basename = os.path.basename(grandparent) if grandparent else ""
        date_str = (
                extract_date(dir_basename)
                or extract_date(parent_basename)
                or extract_date(grandparent_basename)
                or extract_date(dir_path)
        )

        for fname in os.listdir(dir_path):
            lower = fname.lower()
            if not (lower.endswith(".log") or lower.endswith(".txt")):
                continue
            fpath = os.path.join(dir_path, fname)
            if not os.path.isfile(fpath):
                continue

            lines = read_text_lines(fpath)
            if not lines:
                continue

            headers = find_all_subnetwork_headers(lines)
            if not headers:
                continue
            headers.append(len(lines))

            for i in range(len(headers) - 1):
                h, nxt = headers[i], headers[i + 1]
                mo = extract_mo_from_subnetwork_line(lines[h])
                if mo not in ("GUtranCellRelation", "NRCellRelation"):
                    continue

                df = parse_table_slice_from_subnetwork(lines, h, nxt)
                if df is None or df.empty:
                    continue

                df = self._insert_front_columns(df, prepost, date_str)
                # NEW: store file path only for Summary; do not persist it inside DataFrames
                self._source_paths.setdefault(mo, {}).setdefault(prepost, []).append((date_str or "", fpath))
                collected[mo].append(df)

    def loadPrePost(self, input_dir_or_pre: str, post_dir: Optional[str] = None, module_name: Optional[str] = "", market_tag: Optional[str] = "GLOBAL") -> Dict[str, pd.DataFrame]:
        """
        Load Pre/Post either from:
          - Legacy mode: a single parent folder that contains 'Pre' and 'Post' subfolders (post_dir=None), or
          - Dual-input mode: two explicit folders (pre_dir, post_dir) passed in.

        Returns a dict with concatenated GU/NR DataFrames in self.tables.
        """
        collected: Dict[str, List[pd.DataFrame]] = {"GUtranCellRelation": [], "NRCellRelation": []}
        self.pre_folder_found = False
        self.post_folder_found = False

        if post_dir is None:
            # ===== Legacy single-folder mode (original behavior) =====
            input_dir = input_dir_or_pre
            if not os.path.isdir(input_dir):
                raise NotADirectoryError(f"Invalid directory: {input_dir}")

            for entry in os.scandir(input_dir):
                if not entry.is_dir():
                    continue
                prepost = self._detect_prepost(entry.name)
                if not prepost:
                    continue
                if prepost == "Pre":
                    self.pre_folder_found = True
                elif prepost == "Post":
                    self.post_folder_found = True
                self._collect_from_dir(entry.path, prepost, collected)

            if not self.pre_folder_found:
                print(f"{module_name} {market_tag} [INFO] 'Pre' folder not found under: {input_dir}. Returning to GUI.")
            if not self.post_folder_found:
                print(f"{module_name} {market_tag} [INFO] 'Post' folder not found under: {input_dir}. Returning to GUI.")
            if not any(collected.values()):
                print(f"{module_name} {market_tag} [WARNING] No GU/NR tables were loaded from: {input_dir}.")
        else:
            # ===== Dual-input mode: explicit PRE/POST folders =====
            pre_dir = input_dir_or_pre
            if not os.path.isdir(pre_dir):
                raise NotADirectoryError(f"Invalid PRE directory: {pre_dir}")
            if not os.path.isdir(post_dir):
                raise NotADirectoryError(f"Invalid POST directory: {post_dir}")

            self.pre_folder_found = True
            self.post_folder_found = True
            self._collect_from_dir(pre_dir, "Pre", collected)
            self._collect_from_dir(post_dir, "Post", collected)

            if not any(collected.values()):
                print(f"{module_name} {market_tag} [WARNING] No GU/NR tables were loaded from: {pre_dir} and {post_dir}.")

        self.tables = {
            self._table_key_name(base): pd.concat(chunks, ignore_index=True)
            for base, chunks in collected.items() if chunks
        }
        return self.tables

    # ----------------------------- COMPARISON ----------------------------- #
    def comparePrePost(self, freq_before: str, freq_after: str, audit_pre_excel: Optional[str] = None, audit_post_excel: Optional[str] = None, audit_pre_summary_audit_df: Optional[pd.DataFrame] = None, audit_post_summary_audit_df: Optional[pd.DataFrame] = None, module_name: Optional[str] = "", market_tag: Optional[str] = "GLOBAL") -> Dict[str, Dict[str, pd.DataFrame]]:
        """
        Compare Pre/Post relations for GUtranCellRelation and NRCellRelation.

        Parameters:
          freq_before: old SSB frequency (Pre)
          freq_after: new SSB frequency (Post)
          module_name: optional label for logging
          audit_pre_excel: path to PRE ConfigurationAudit Excel (SummaryAudit sheet)
          audit_post_excel: path to POST ConfigurationAudit Excel (SummaryAudit sheet)

        Note:
          audit_post_excel is used to exclude nodes that did not complete retuning.
          Both audit_pre_excel and audit_post_excel are stored in the instance so that
          save_outputs_excel can build the SummaryAuditComparisson sheet.
        """

        # NEW: store audit paths in the instance so they can be used later in save_outputs_excel
        def _resolve_audit_excel_path(p: Optional[str]) -> Optional[str]:
            # Accept only a non-empty existing .xlsx file path; otherwise return None.
            if not p:
                return None
            try:
                p_long = to_long_path(p)
            except Exception:
                p_long = p
            if os.path.isfile(p_long) and str(p_long).lower().endswith(".xlsx") and os.path.getsize(p_long) > 0:
                return p_long
            return None

        self.audit_pre_excel = _resolve_audit_excel_path(audit_pre_excel)
        self.audit_post_excel = _resolve_audit_excel_path(audit_post_excel)

        def _fast_load_summaryaudit_df(path: Optional[str], label: str) -> Optional[pd.DataFrame]:
            if not path:
                return None
            try:
                x_path = to_long_path(path)
            except Exception:
                x_path = path
            if not os.path.isfile(x_path):
                return None
            try:
                wanted = {"Category", "SubCategory", "Metric", "Value", "ExtraInfo"}
                return pd.read_excel(x_path, sheet_name="SummaryAudit", usecols=lambda c: c in wanted, engine="openpyxl")
            except Exception:
                return None

        # Cache SummaryAudit DataFrames (avoid reading the same Excel multiple times)
        pre_summary_df = audit_pre_summary_audit_df if isinstance(audit_pre_summary_audit_df, pd.DataFrame) else _fast_load_summaryaudit_df(audit_pre_excel, "PRE")
        post_summary_df = audit_post_summary_audit_df if isinstance(audit_post_summary_audit_df, pd.DataFrame) else _fast_load_summaryaudit_df(audit_post_excel, "POST")

        self.audit_pre_summary_audit_df = pre_summary_df
        self.audit_post_summary_audit_df = post_summary_df

        if not self.tables:
            # Soft fail: do not raise, just inform and return empty results
            print(f"{module_name} {market_tag} [WARNING] No tables loaded. Skipping comparison (likely missing Pre/Post folders).")
            return {}

        # NEW: load node numeric identifiers from nodes with Retuning frpm POST Configuration Audit (SummaryAudit sheet)
        nodes_id_pre, nodes_name_pre = load_nodes_names_and_id_from_summary_audit(post_summary_df if post_summary_df is not None else audit_post_excel, stage="Pre", module_name=module_name)
        nodes_id_post, nodes_name_post = load_nodes_names_and_id_from_summary_audit(post_summary_df if post_summary_df is not None else audit_post_excel, stage="Post", module_name=module_name)

        results: Dict[str, Dict[str, pd.DataFrame]] = {}

        for table_name, df_all in self.tables.items():
            if df_all.empty:
                continue

            freq_col = detect_freq_column(table_name, list(df_all.columns))
            if not freq_col:
                print(f"{module_name} {market_tag} [WARNING] No frequency column detected in {table_name}. Adjust mapping if needed.")
                continue

            key_cols = detect_key_columns(table_name, list(df_all.columns), freq_col)
            if not key_cols:
                print(f"{module_name} {market_tag} [WARNING] No key column detected in {table_name}.")
                continue

            # Forzar claves estables si existen
            if table_name == "GUtranCellRelation":
                forced = [c for c in ["NodeId", "EUtranCellFDDId", "GUtranCellRelationId"] if c in df_all.columns]
                if forced:
                    key_cols = forced
            elif table_name == "NRCellRelation":
                forced = [c for c in ["NodeId", "NRCellCUId", "NRCellRelationId"] if c in df_all.columns]
                if forced:
                    key_cols = forced

            pre_df_full = select_latest_by_date(df_all, "Pre")
            post_df_full = select_latest_by_date(df_all, "Post")

            if pre_df_full.empty and post_df_full.empty:
                continue

            # Pick representative source file for Summary from internal paths (no DF column)
            def pick_src(tbl: str, side: str, target_date: str) -> str:
                pool = self._source_paths.get(tbl, {}).get(side, [])
                # Prefer exact date match (latest set), else first available
                for d, p in pool:
                    if d == target_date:
                        return p
                return pool[0][1] if pool else ""

            pre_date = pre_df_full["Date"].max() if not pre_df_full.empty and "Date" in pre_df_full.columns else ""
            post_date = post_df_full["Date"].max() if not post_df_full.empty and "Date" in post_df_full.columns else ""
            pre_source_file = pick_src(table_name, "Pre", pre_date)
            post_source_file = pick_src(table_name, "Post", post_date)

            pre_norm = normalize_df(pre_df_full)
            post_norm = normalize_df(post_df_full)

            pre_idx = make_index_by_keys(pre_norm, key_cols)
            post_idx = make_index_by_keys(post_norm, key_cols)

            pre_keys, post_keys = set(pre_idx.index), set(post_idx.index)
            common_idx = sorted(pre_keys & post_keys)

            new_in_post = post_idx.loc[sorted(post_keys - pre_keys)].copy()
            missing_in_post = pre_idx.loc[sorted(pre_keys - post_keys)].copy()

            pre_common = pre_idx.loc[common_idx]
            post_common = post_idx.loc[common_idx]

            def slim(df: pd.DataFrame, keep_cols: List[str]) -> pd.DataFrame:
                cols = ["Pre/Post", "Date"] + list(dict.fromkeys(keep_cols))
                cols = [c for c in cols if c in df.columns]
                return df[cols].copy()

            pre_slim = slim(pre_common, key_cols + [freq_col])
            post_slim = slim(post_common, key_cols + [freq_col])

            # Freq base
            if table_name == "NRCellRelation":
                pre_freq_base = extract_nr_freq_base(pre_slim.get(freq_col, pd.Series("", index=pre_slim.index)))
                post_freq_base = extract_nr_freq_base(post_slim.get(freq_col, pd.Series("", index=post_slim.index)))
            else:
                pre_freq_base = extract_gu_freq_base(pre_slim.get(freq_col, pd.Series("", index=pre_slim.index)))
                post_freq_base = extract_gu_freq_base(post_slim.get(freq_col, pd.Series("", index=post_slim.index)))

            fb, fa = str(freq_before).strip(), str(freq_after).strip()
            pre_has_before = (pre_freq_base == fb)
            pre_has_after = (pre_freq_base == fa)
            post_is_after = (post_freq_base == fa)
            freq_rule_mask = (pre_has_before & (~post_is_after)) | (pre_has_after & (~post_is_after))

            exclude_cols = {"Pre/Post", "Date", freq_col} | set(key_cols)
            shared_cols = [c for c in pre_common.columns if c in post_common.columns and c not in exclude_cols]

            # NEW: ignore timeOfCreation differences for GU discrepancies
            if table_name == "GUtranCellRelation":
                shared_cols = [c for c in shared_cols if c != "timeOfCreation"]

            any_diff_mask = pd.Series(False, index=pre_common.index)
            diff_cols_per_row = {k: [] for k in pre_common.index}
            for c in shared_cols:
                diffs = (pre_common[c] != post_common[c]).reindex(pre_common.index, fill_value=False)
                any_diff_mask = any_diff_mask | diffs
                for k in pre_common.index[diffs]:
                    diff_cols_per_row[k].append(c)

            # NEW: optionally exclude discrepancies for relations whose destination nodes did not complete the retuning
            rel_series = None
            pattern_nodes = None
            if nodes_id_pre and table_name in ("GUtranCellRelation", "NRCellRelation"):
                relation_col = "GUtranCellRelationId" if table_name == "GUtranCellRelation" else "NRCellRelationId"

                # Ensure the relation column exists either in PRE or POST
                if relation_col in post_common.columns or relation_col in pre_common.columns:
                    # Choose POST table if available (latest data), otherwise PRE
                    src_rel_df = post_common if relation_col in post_common.columns else pre_common
                    # Convert the relation column to a clean string series
                    rel_series = src_rel_df[relation_col].reindex(pre_common.index).astype(str).fillna("")
                    try:
                        # Build a substring-matching pattern for all numeric node identifiers
                        pattern_nodes = "|".join(re.escape(x) for x in sorted(nodes_id_pre))
                        # Build the skip mask: all rows whose relation contains a non-retuned node id
                        skip_mask = rel_series.str.contains(pattern_nodes, regex=True, na=False)
                        # Remove those rows from discrepancy masks (parameter and frequency differences)
                        any_diff_mask = any_diff_mask & ~skip_mask
                        freq_rule_mask = freq_rule_mask & ~skip_mask
                    except re.error:
                        # If regex construction fails (rare), do not modify any masks
                        pass


                    # NEW: classify frequency discrepancies into SSB-Post vs Unknown based on destination target ids
                    def _extract_kv_from_ref(ref_value: object, key: str) -> str:
                        text = str(ref_value or "")
                        m = re.search(rf"{re.escape(key)}=([^,]+)", text)
                        return m.group(1).strip() if m else ""

                    def _detect_gnodeb_target(ext_id: object) -> str:
                        val = str(ext_id) if ext_id is not None else ""
                        if nodes_id_pre and any(n in val for n in nodes_id_pre):
                            return "SSB-Pre"
                        if nodes_id_post and any(n in val for n in nodes_id_post):
                            return "SSB-Post"
                        return "Unknown"

                    gnodeb_target_series = pd.Series("Unknown", index=pre_common.index)
                    ext_gnb_series = pd.Series("", index=pre_common.index)
                    ext_cell_series = pd.Series("", index=pre_common.index)

                    if table_name == "NRCellRelation":
                        ref_col = None
                        for cand in ["nRCellRef", "NRCellRef", "neighborCellRef"]:
                            if cand in post_common.columns:
                                ref_col = cand
                                break
                            if cand in pre_common.columns:
                                ref_col = cand
                                break

                        if ref_col:
                            ref_series = (post_common[ref_col] if ref_col in post_common.columns else pre_common[ref_col]).reindex(pre_common.index)
                            ext_gnb_series = ref_series.map(lambda v: _extract_kv_from_ref(v, "ExternalGNBCUCPFunction"))
                            ext_cell_series = ref_series.map(lambda v: _extract_kv_from_ref(v, "ExternalNRCellCU"))
                            gnodeb_target_series = ext_gnb_series.map(_detect_gnodeb_target)

                    elif table_name == "GUtranCellRelation":
                        ref_col = None
                        for cand in ["neighborCellRef", "nCellRef", "NCellRef"]:
                            if cand in post_common.columns:
                                ref_col = cand
                                break
                            if cand in pre_common.columns:
                                ref_col = cand
                                break

                        if ref_col:
                            ref_series = (post_common[ref_col] if ref_col in post_common.columns else pre_common[ref_col]).reindex(pre_common.index)
                            ext_gnb_series = ref_series.map(lambda v: _extract_kv_from_ref(v, "ExternalGNodeBFunction"))
                            ext_cell_series = ref_series.map(lambda v: _extract_kv_from_ref(v, "ExternalGUtranCell"))
                            gnodeb_target_series = ext_gnb_series.map(_detect_gnodeb_target)

            # Exclude frequency-only mismatches for relations whose destination is still SSB-Pre (no retuning)
            try:
                freq_rule_mask = freq_rule_mask & (gnodeb_target_series.astype(str).str.strip() != "SSB-Pre")
            except Exception:
                pass

            combined_mask = (freq_rule_mask | any_diff_mask).reindex(pre_common.index, fill_value=False)

            discrepancy_keys = [k for k, m in zip(pre_common.index, combined_mask) if m and k in set(common_idx)]

            # Build discrepancies
            def desired_key_order(tbl: str) -> list:
                if tbl == "GUtranCellRelation":
                    return ["NodeId", "EUtranCellFDDId", "GUtranCellRelationId"]
                if tbl == "NRCellRelation":
                    return ["NodeId", "NRCellCUId", "NRCellRelationId"]
                return []

            def reorder_cols(df: pd.DataFrame, tbl: str) -> pd.DataFrame:
                if df is None or df.empty:
                    return df
                front = ["Date_Pre", "Date_Post", "Freq_Pre", "Freq_Post"]
                keys = [c for c in desired_key_order(tbl) if c in df.columns]
                seen = set(front + keys)
                rest = [c for c in df.columns if c not in seen]
                return df[[*(c for c in front if c in df.columns), *keys, *rest]]

            rows = []
            for k in discrepancy_keys:
                row = {}
                for c in key_cols:
                    row[c] = pre_common.loc[k, c] if c in pre_common.columns else ""
                row["Date_Pre"] = pre_slim.loc[k, "Date"] if "Date" in pre_slim.columns else ""
                row["Date_Post"] = post_slim.loc[k, "Date"] if "Date" in post_slim.columns else ""

                # For NR, store only the base frequency (e.g. 648672) instead of the full ref string
                if table_name == "NRCellRelation":
                    row["Freq_Pre"] = pre_freq_base.get(k, "")
                    row["Freq_Post"] = post_freq_base.get(k, "")
                else:
                    row["Freq_Pre"] = (
                        pre_slim.get(freq_col, pd.Series("", index=pre_slim.index)).loc[k]
                        if k in pre_slim.index else ""
                    )
                    row["Freq_Post"] = (
                        post_slim.get(freq_col, pd.Series("", index=post_slim.index)).loc[k]
                        if k in post_slim.index else ""
                    )

                required_cols = (
                    ["NodeId", "EUtranCellFDDId", "GUtranFreqRelationId", "GUtranCellRelationId"]
                    if table_name == "GUtranCellRelation"
                    else ["NodeId", "NRCellCUId", "NRCellRelationId"]
                )

                for rc in required_cols:
                    val = ""
                    if rc in post_common.columns:
                        val = post_common.loc[k, rc]
                    elif rc in pre_common.columns:
                        val = pre_common.loc[k, rc]
                    row[rc] = val

                if table_name == "NRCellRelation":
                    row["ExternalGNBCUCPFunction"] = ext_gnb_series.loc[k] if k in ext_gnb_series.index else ""
                    row["ExternalNRCellCU"] = ext_cell_series.loc[k] if k in ext_cell_series.index else ""
                    row["GNodeB_SSB_Target"] = gnodeb_target_series.loc[k] if k in gnodeb_target_series.index else "Unknown"
                elif table_name == "GUtranCellRelation":
                    row["ExternalGNodeBFunction"] = ext_gnb_series.loc[k] if k in ext_gnb_series.index else ""
                    row["ExternalGUtranCell"] = ext_cell_series.loc[k] if k in ext_cell_series.index else ""
                    row["GNodeB_SSB_Target"] = gnodeb_target_series.loc[k] if k in gnodeb_target_series.index else "Unknown"

                difflist = diff_cols_per_row.get(k, [])

                # NEW: when there is no parameter difference but the frequency rule
                #      says this relation is inconsistent (SSB not updated), add
                #      a descriptive text in DiffColumns.
                is_freq_only_mismatch = False
                try:
                    is_freq_only_mismatch = bool(freq_rule_mask.loc[k]) and not difflist
                except KeyError:
                    is_freq_only_mismatch = False

                if is_freq_only_mismatch:
                    target_val = "Unknown"
                    try:
                        target_val = str(gnodeb_target_series.loc[k] if k in gnodeb_target_series.index else "Unknown").strip()
                    except Exception:
                        target_val = "Unknown"

                    if target_val == "SSB-Post":
                        row["DiffColumns"] = "SSB Post-Retuning keeps equal than SSB Pre-Retuning"
                    else:
                        row["DiffColumns"] = "SSB Target Unknown"
                else:
                    row["DiffColumns"] = ", ".join(sorted(difflist))

                for c in difflist:
                    row[f"{c}_Pre"] = pre_common.loc[k, c]
                    row[f"{c}_Post"] = post_common.loc[k, c]

                rows.append(row)

            discrepancies = reorder_cols(pd.DataFrame(rows), table_name)

            if not new_in_post.empty:
                for col in new_in_post.columns:
                    new_in_post[col] = new_in_post[col].astype(str)
            if not missing_in_post.empty:
                for col in missing_in_post.columns:
                    missing_in_post[col] = missing_in_post[col].astype(str)

            # --- light construction for new/missing tables (only keys + Freq_Pre/Freq_Post) ---
            def with_freq_pair(df_src: pd.DataFrame, tbl: str, kind: str) -> pd.DataFrame:
                """
                Build a light table for _new / _missing:
                  - Use only key columns (plus NodeId if present).
                  - Compute base frequency from freq_col.
                  - For 'new':   Freq_Pre = ""        , Freq_Post = base
                  - For 'missing': Freq_Pre = base    , Freq_Post = ""
                  - Do not drag all relation columns; keep them only in all_relations.
                """
                if df_src is None or df_src.empty:
                    return df_src

                df_src = df_src.copy()
                for col in df_src.columns:
                    df_src[col] = df_src[col].astype(str)

                # Base frequency from main freq_col
                if tbl == "NRCellRelation":
                    base = extract_nr_freq_base(df_src.get(freq_col, pd.Series("", index=df_src.index)))
                else:
                    base = extract_gu_freq_base(df_src.get(freq_col, pd.Series("", index=df_src.index)))

                # Keep only key columns (and NodeId if not already included)
                keep_cols: List[str] = []
                if "NodeId" in df_src.columns:
                    keep_cols.append("NodeId")
                for c in key_cols:
                    if c in df_src.columns and c not in keep_cols:
                        keep_cols.append(c)

                df_tmp = df_src[keep_cols].copy()

                if kind == "new":
                    df_tmp["Freq_Pre"] = ""
                    df_tmp["Freq_Post"] = base
                elif kind == "missing":
                    df_tmp["Freq_Pre"] = base
                    df_tmp["Freq_Post"] = ""
                else:
                    df_tmp["Freq_Pre"] = ""
                    df_tmp["Freq_Post"] = ""

                return df_tmp

            new_in_post_clean = with_freq_pair(new_in_post, table_name, kind="new")
            missing_in_post_clean = with_freq_pair(missing_in_post, table_name, kind="missing")

            # NEW: optional frequency-based filter for _disc / _new / _missing tables
            discrepancies = self._filter_rows_by_freq_list(discrepancies)
            new_in_post_clean = self._filter_rows_by_freq_list(new_in_post_clean)
            missing_in_post_clean = self._filter_rows_by_freq_list(missing_in_post_clean)

            # Pair stats
            freq_diff_series = freq_rule_mask.reindex(pre_common.index).astype(bool)
            pair_stats = pd.DataFrame(
                {
                    "Freq_Pre": pre_freq_base.reindex(pre_common.index).fillna("").replace("", "<empty>"),
                    "Freq_Post": post_freq_base.reindex(pre_common.index).fillna("").replace("", "<empty>"),
                    "ParamDiff": any_diff_mask.reindex(pre_common.index).astype(bool),
                    "FreqDiff": freq_diff_series,
                    "FreqDiff_SSBPost": freq_diff_series & (gnodeb_target_series.astype(str).str.strip() == "SSB-Post"),
                    "FreqDiff_Unknown": freq_diff_series & (gnodeb_target_series.astype(str).str.strip() == "Unknown"),

                },
                index=pre_common.index,
            )

            # all_relations (merge último PRE/POST, manteniendo Freq_Pre/Freq_Post)
            pre_latest = pre_norm.copy()
            post_latest = post_norm.copy()
            if table_name == "NRCellRelation":
                pre_fb = extract_nr_freq_base(pre_latest.get(freq_col, pd.Series("", index=pre_latest.index)))
                post_fb = extract_nr_freq_base(post_latest.get(freq_col, pd.Series("", index=post_latest.index)))
            else:
                pre_fb = extract_gu_freq_base(pre_latest.get(freq_col, pd.Series("", index=pre_latest.index)))
                post_fb = extract_gu_freq_base(post_latest.get(freq_col, pd.Series("", index=post_latest.index)))
            pre_latest = pre_latest.assign(Freq_Pre=pre_fb.replace("", "<empty>"))
            post_latest = post_latest.assign(Freq_Post=post_fb.replace("", "<empty>"))

            def keys_first(df: pd.DataFrame) -> pd.DataFrame:
                if df is None or df.empty:
                    return df
                ko = [c for c in key_cols if c in df.columns]
                rest = [c for c in df.columns if c not in ko]
                return df[ko + rest]

            pre_keep = keys_first(pre_latest.drop(columns=["Pre/Post", "Date"], errors="ignore"))
            post_keep = keys_first(post_latest.drop(columns=["Pre/Post", "Date"], errors="ignore"))

            merged_all = pd.merge(pre_keep, post_keep, on=key_cols, how="outer", suffixes=("_PreSide", "_PostSide"))
            all_relations = merged_all[key_cols].copy()
            all_relations["Freq_Pre"] = merged_all.get("Freq_Pre", "")
            all_relations["Freq_Post"] = merged_all.get("Freq_Post", "")

            for col in set(pre_keep.columns) | set(post_keep.columns):
                if col in key_cols or col in ("Freq_Pre", "Freq_Post"):
                    continue

                pre_col = f"{col}_PreSide"
                post_col = f"{col}_PostSide"

                if post_col in merged_all.columns and pre_col in merged_all.columns:
                    # Prefer POST value only if it is not empty/NaN;
                    # otherwise fall back to PRE value.
                    post_series = merged_all[post_col]
                    as_str = post_series.astype(str).str.strip().str.lower()
                    is_empty = as_str.isin(("", "nan"))  # treat NaN and empty as "no value"

                    all_relations[col] = post_series.where(
                        ~is_empty,
                        merged_all[pre_col],
                    )

                elif post_col in merged_all.columns:
                    all_relations[col] = merged_all[post_col]

                elif pre_col in merged_all.columns:
                    all_relations[col] = merged_all[pre_col]

                elif col in merged_all.columns:
                    all_relations[col] = merged_all[col]

            results[table_name] = {
                "discrepancies": discrepancies.reset_index(drop=True),
                "new_in_post": new_in_post_clean.reset_index(drop=True),
                "missing_in_post": missing_in_post_clean.reset_index(drop=True),
                "pair_stats": pair_stats.reset_index(drop=True),
                "all_relations": all_relations.reset_index(drop=True),
                "meta": {
                    "key_cols": key_cols,
                    "freq_col": freq_col,
                    "pre_rows": int(pre_df_full.shape[0]),
                    "post_rows": int(post_df_full.shape[0]),
                    "pre_source_file": pre_source_file,  # NEW
                    "post_source_file": post_source_file,  # NEW
                },
            }

            print(f"\n{module_name} {market_tag} [INFO] === {table_name} ===")
            print(f"{module_name} {market_tag} [INFO] Key: {key_cols} | Freq column: {freq_col}")

            freq_disc_count = 0
            ssb_unknown_count = 0
            param_disc_count = 0
            try:
                if not pair_stats.empty:
                    if "ParamDiff" in pair_stats.columns:
                        param_disc_count = int(pd.to_numeric(pair_stats["ParamDiff"], errors="coerce").fillna(0).astype(int).sum())
                    if "FreqDiff_SSBPost" in pair_stats.columns:
                        freq_disc_count = int(pd.to_numeric(pair_stats["FreqDiff_SSBPost"], errors="coerce").fillna(0).astype(int).sum())
                    elif "FreqDiff" in pair_stats.columns:
                        freq_disc_count = int(pd.to_numeric(pair_stats["FreqDiff"], errors="coerce").fillna(0).astype(int).sum())
                    if "FreqDiff_Unknown" in pair_stats.columns:
                        ssb_unknown_count = int(pd.to_numeric(pair_stats["FreqDiff_Unknown"], errors="coerce").fillna(0).astype(int).sum())
            except Exception:
                freq_disc_count = 0
                ssb_unknown_count = 0
                param_disc_count = 0

            relations_with_discrepancies = int(len(discrepancies))
            total_discrepancies = int(param_disc_count) + int(freq_disc_count) + int(ssb_unknown_count)

            print(f"{module_name} {market_tag} [INFO] - Total Discrepancies (one relation can have more that one type of discrepancy): {total_discrepancies}")
            print(f"{module_name} {market_tag} [INFO]   - Param Discrepancies: {param_disc_count}")
            print(f"{module_name} {market_tag} [INFO]   - Frequency Discrepancies (SSB-Post): {freq_disc_count}")
            print(f"{module_name} {market_tag} [INFO]   - SSB Unknown: {ssb_unknown_count}")
            print(f"{module_name} {market_tag} [INFO] - Relations with Discrepancies (uniques): {relations_with_discrepancies}")
            print(f"{module_name} {market_tag} [INFO] - New Relations in Post: {len(new_in_post_clean)}")
            print(f"{module_name} {market_tag} [INFO] - Missing Relations in Post: {len(missing_in_post_clean)}")

            # Print the relation names that match the pattern and will be excluded
            if rel_series is not None and not rel_series.empty and pattern_nodes:
                to_skip_relations = rel_series[rel_series.str.contains(pattern_nodes, regex=True, na=False)]
                if not to_skip_relations.empty:
                    skipped_count = len(to_skip_relations)
                    print(f"{module_name} {market_tag} [INFO] - Relations skipped due to destination node being in the no-retuning buffer ({table_name}): {skipped_count} ")
                    # Below line show the list of all skipped relations (it can be huge, better keep below line commented).
                    # print(f"{module_name} {market_tag} [INFO] - Relations skipped List: {sorted(to_skip_relations.unique())}")

        return results

    # ----------------------------- SUMMARY AUDIT COMPARISSON ----------------------------- #
    def summaryaudit_comparison(self, module_name: str = "", market_tag: str = "GLOBL") -> Optional[pd.DataFrame]:
        """
        Build a comparison DataFrame from PRE and POST ConfigurationAudit SummaryAudit sheets.

        - Loads 'SummaryAudit' from both PRE and POST audit Excels.
        - Drops the 'ExtraInfo' column if present.
        - Renames 'Value' to 'Value_Pre' in PRE and to 'Value_Post' in POST.
        - Merges both on all common columns except the value columns.
        - Row order is driven by the PRE sheet (left-merge), so the new sheet
          keeps exactly the same row ordering as the PRE SummaryAudit.
        """
        pre_path = self.audit_pre_excel
        post_path = self.audit_post_excel

        pre_cached = self.audit_pre_summary_audit_df if isinstance(self.audit_pre_summary_audit_df, pd.DataFrame) else None
        post_cached = self.audit_post_summary_audit_df if isinstance(self.audit_post_summary_audit_df, pd.DataFrame) else None

        if (pre_cached is None and not pre_path) and (post_cached is None and not post_path):
            return None

        def _load_summary(path: str, label: str) -> Optional[pd.DataFrame]:
            try:
                x_path = to_long_path(path)
            except Exception:
                x_path = path

            if not os.path.isfile(x_path):
                print(f"{module_name} {market_tag} [WARNING] {label} SummaryAudit Excel not found: '{path}'. Skipping SummaryAuditComparisson for this side.")
                return None

            try:
                wanted = {"Category", "SubCategory", "Metric", "Value"}
                df = pd.read_excel(x_path, sheet_name="SummaryAudit", usecols=lambda c: c in wanted, engine="openpyxl")
            except Exception as e:
                print(f"{module_name} {market_tag} [WARNING] Could not read 'SummaryAudit' sheet from {label} audit Excel '{path}': {e}.")
                return None

            if "ExtraInfo" in df.columns:
                df = df.drop(columns=["ExtraInfo"])
            if "Value" not in df.columns:
                df["Value"] = pd.NA

            return df

        print(f"{module_name} {market_tag} [INFO] Loading ConfigurationAudits to extract SummaryAudit sheets...")
        pre_df = pre_cached.copy() if pre_cached is not None else (_load_summary(pre_path, "PRE") if pre_path else None)
        post_df = post_cached.copy() if post_cached is not None else (_load_summary(post_path, "POST") if post_path else None)

        # Keep ExtraInfo in SummaryAudit for other checks, but exclude it from SummaryAuditComparisson
        cmp_cols = ["Category", "SubCategory", "Metric", "Value"]

        if isinstance(pre_df, pd.DataFrame):
            if "Value" not in pre_df.columns:
                pre_df["Value"] = pd.NA
            pre_df = pre_df.loc[:, [c for c in cmp_cols if c in pre_df.columns]].copy()

        if isinstance(post_df, pd.DataFrame):
            if "Value" not in post_df.columns:
                post_df["Value"] = pd.NA
            post_df = post_df.loc[:, [c for c in cmp_cols if c in post_df.columns]].copy()

        if pre_df is None and post_df is None:
            return None
        if pre_df is None:
            # Only POST available: just rename its Value as Value_Post
            post_df = post_df.copy()
            post_df = post_df.rename(columns={"Value": "Value_Post"})
            print(f"{module_name} {market_tag} [INFO] Using only 'SummaryAudit' from 'ConfigurationAudit POST' SummaryAudit to generate 'SummaryAuditComparisson' sheet...")

            return post_df
        if post_df is None:
            # Only PRE available: just rename its Value as Value_Pre
            pre_df = pre_df.copy()
            pre_df = pre_df.rename(columns={"Value": "Value_Pre"})
            print(f"{module_name} {market_tag} [INFO] Using only 'SummaryAudit' from 'ConfigurationAudit PRE' SummaryAudit to generate 'SummaryAuditComparisson' sheet...")
            return pre_df

        # NEW: copy before renaming to avoid side effects
        pre_df = pre_df.copy()
        post_df = post_df.copy()

        pre_df = pre_df.rename(columns={"Value": "Value_Pre"})
        post_df = post_df.rename(columns={"Value": "Value_Post"})

        # Common columns to merge on (all shared columns except the value columns)
        common_cols = [
            c for c in pre_df.columns
            if c in post_df.columns and c not in ("Value_Pre", "Value_Post")
        ]

        if not common_cols:
            # If there are no common columns, just concatenate with an extra column indicating source
            pre_df["Source"] = "PRE"
            post_df["Source"] = "POST"
            merged = pd.concat([pre_df, post_df], ignore_index=True)
        else:
            # NEW: perform a LEFT merge using PRE as reference to preserve row order
            #      This guarantees that SummaryAuditComparisson keeps the exact PRE ordering.
            merged = pd.merge(
                pre_df,
                post_df[common_cols + ["Value_Post"]],
                on=common_cols,
                how="left",
                sort=False,
            )

        print(f"{module_name} {market_tag} [INFO] Using PRE and POST 'SummaryAudit' sheets from 'ConfigurationAudit' module to generate 'SummaryAuditComparisson' sheet....")

        # Add numeric difference column at the end: Value_Pre - Value_Post
        try:
            vpre = pd.to_numeric(merged.get("Value_Pre"), errors="coerce")
            vpost = pd.to_numeric(merged.get("Value_Post"), errors="coerce")
            merged["Value_Diff"] = vpre - vpost
        except Exception:
            merged["Value_Diff"] = pd.NA

        try:
            cols = [c for c in merged.columns if c != "Value_Diff"] + ["Value_Diff"]
            merged = merged[cols]
        except Exception:
            pass

        return merged

    # ----------------------------- OUTPUT TO EXCEL ----------------------------- #
    def save_outputs_excel(self, output_dir: str, results: Optional[Dict[str, Dict[str, pd.DataFrame]]] = None, versioned_suffix: Optional[str] = None, module_name: str = "", market_tag: str = "GLOBAL", fast_excel_export: bool = False, fast_excel_autofit_rows: int = 200, fast_excel_autofit_max_width: int = 60) -> None:
        import os
        import tempfile
        import shutil

        os.makedirs(output_dir, exist_ok=True)
        suf = f"_{versioned_suffix}" if versioned_suffix else ""

        # Path for output files (final destination)
        excel_cell_relation = os.path.join(output_dir, f"CellRelation{suf}.xlsx")
        excel_cc_cell_relation = os.path.join(output_dir, f"ConsistencyChecks_CellRelation{suf}.xlsx")

        # Convert paths to long-path form on Windows to avoid MAX_PATH issues
        excel_cell_relation_long = to_long_path(excel_cell_relation)
        excel_cc_cell_relation_long = to_long_path(excel_cc_cell_relation)

        # -------------------------------------------------------------------
        #  Write into SYSTEM TEMP first (avoid OneDrive sync issues), then move to output_dir
        # -------------------------------------------------------------------
        def _make_temp_xlsx_path(final_xlsx_path: str) -> tuple:
            """
            Create a temp dir and a temp xlsx path. Prefer system temp to avoid OneDrive sync during write.
            Returns (tmp_dir, tmp_xlsx_path).
            """
            tmp_dir = tempfile.mkdtemp(prefix="SSB_RA_")
            tmp_name = os.path.basename(final_xlsx_path)
            tmp_xlsx = os.path.join(tmp_dir, tmp_name)
            return tmp_dir, tmp_xlsx

        def _move_into_place(tmp_path: str, final_path: str) -> None:
            """
            Move temp file into final destination.
            - Try atomic replace if possible.
            - Fallback to shutil.move (handles cross-device).
            """
            try:
                os.replace(tmp_path, final_path)
            except Exception:
                shutil.move(tmp_path, final_path)

        def _pick_excel_engine() -> str:
            if not fast_excel_export:
                return "openpyxl"
            try:
                import xlsxwriter  # noqa: F401
                return "xlsxwriter"
            except Exception:
                return "openpyxl"

        excel_engine = _pick_excel_engine()

        tmp_dir_cc, tmp_excel_cc = _make_temp_xlsx_path(excel_cc_cell_relation_long)
        tmp_dir_cell, tmp_excel_cell = _make_temp_xlsx_path(excel_cell_relation_long)

        tmp_excel_cc_long = to_long_path(tmp_excel_cc)
        tmp_excel_cell_long = to_long_path(tmp_excel_cell)

        try:
            # -------------------------------------------------------------------
            #  Write ConsistencyChecks_CellRelations.xlsx (TEMP)
            # -------------------------------------------------------------------
            print(f"{module_name} {market_tag} [INFO] Saving {pretty_path(excel_cc_cell_relation)} (engine={excel_engine}, temp -> final)...")

            written_sheet_dfs: Dict[str, pd.DataFrame] = {}

            if excel_engine == "xlsxwriter":
                writer_ctx = pd.ExcelWriter(tmp_excel_cc_long, engine="xlsxwriter", engine_kwargs={"options": {"strings_to_urls": False, "strings_to_numbers": False}})
            else:
                writer_ctx = pd.ExcelWriter(tmp_excel_cc_long, engine="openpyxl")

            with writer_ctx as writer:
                # Summary
                summary_rows = []
                if results:
                    for name, bucket in results.items():
                        meta = bucket.get("meta", {})
                        pair_stats = bucket.get("pair_stats", pd.DataFrame())
                        params_disc = int(pair_stats["ParamDiff"].sum()) if not pair_stats.empty and "ParamDiff" in pair_stats.columns else 0
                        ssb_unknown = int(pair_stats["FreqDiff_Unknown"].sum()) if not pair_stats.empty and "FreqDiff_Unknown" in pair_stats.columns else 0
                        freq_disc = int(pair_stats["FreqDiff_SSBPost"].sum()) if not pair_stats.empty and "FreqDiff_SSBPost" in pair_stats.columns else 0

                        summary_rows.append({
                            "Table": name,
                            "KeyColumns": ", ".join(meta.get("key_cols", [])),
                            "FreqColumn": meta.get("freq_col", "N/A"),
                            "Relations_Pre": meta.get("pre_rows", 0),
                            "Relations_Post": meta.get("post_rows", 0),
                            "Parameters_Discrepancies": params_disc,
                            "Freq_Discrepancies": freq_disc,
                            "SSB_Unknown": ssb_unknown,
                            "New_Relations": len(bucket.get("new_in_post", pd.DataFrame())),
                            "Missing_Relations": len(bucket.get("missing_in_post", pd.DataFrame())),
                            "SourceFile_Pre": pretty_path(meta.get("pre_source_file", "")),
                            "SourceFile_Post": pretty_path(meta.get("post_source_file", "")),
                        })

                summary_df = pd.DataFrame(summary_rows) if summary_rows else pd.DataFrame(columns=["Table", "KeyColumns", "FreqColumn", "Relations_Pre", "Relations_Post", "Parameters_Discrepancies", "Freq_Discrepancies", "SSB_Unknown", "New_Relations", "Missing_Relations", "SourceFile_Pre", "SourceFile_Post"])
                summary_df.to_excel(writer, sheet_name="Summary", index=False)
                written_sheet_dfs["Summary"] = summary_df

                # NEW: add SummaryAuditComparisson sheet if PRE/POST ConfigurationAudit SummaryAudit are available
                comparison_df = self.summaryaudit_comparison(module_name=module_name, market_tag=market_tag)
                if comparison_df is not None and not comparison_df.empty:
                    comparison_df.to_excel(writer, sheet_name="SummaryAuditComparisson", index=False)
                    written_sheet_dfs["SummaryAuditComparisson"] = comparison_df

                # Summary_CellRelation
                detailed_rows = []
                if results:
                    for name, bucket in results.items():
                        meta = bucket.get("meta", {})
                        pair_stats = bucket.get("pair_stats", pd.DataFrame())
                        new_df = bucket.get("new_in_post", pd.DataFrame())
                        miss_df = bucket.get("missing_in_post", pd.DataFrame())

                        def count_side(side: str) -> Dict[str, int]:
                            tbl = select_latest_by_date(self.tables.get(name, pd.DataFrame()), side)
                            if tbl is None or tbl.empty:
                                return {}
                            if name == "NRCellRelation":
                                col = next((c for c in tbl.columns if c.lower() == "nrfreqrelationref"), None)
                                if col:
                                    ser = tbl[col].astype(str).str.extract(r"NRFreqRelation\s*=\s*(\d+)", expand=False).fillna("")
                                else:
                                    ser = (tbl[meta.get("freq_col")].astype(str).str.split("-", n=1).str[0] if meta.get("freq_col") in tbl.columns else pd.Series("", index=tbl.index))
                            else:
                                ser = (tbl[meta.get("freq_col")].astype(str).str.split("-", n=1).str[0] if meta.get("freq_col") in tbl.columns else pd.Series("", index=tbl.index))
                            return ser.fillna("").replace("", "<empty>").value_counts().to_dict()

                        pre_counts = count_side("Pre")
                        post_counts = count_side("Post")

                        if not pair_stats.empty:
                            grp = pair_stats.groupby(["Freq_Pre", "Freq_Post"], dropna=False)
                            params_by_pair = grp["ParamDiff"].sum().astype(int).to_dict() if "ParamDiff" in pair_stats.columns else {}
                            freq_by_pair = (grp["FreqDiff_SSBPost"].sum().astype(int).to_dict() if "FreqDiff_SSBPost" in pair_stats.columns else {})
                            unknown_by_pair = (grp["FreqDiff_Unknown"].sum().astype(int).to_dict() if "FreqDiff_Unknown" in pair_stats.columns else {})
                            pairs_present = set(grp.size().index.tolist())
                        else:
                            params_by_pair, freq_by_pair, pairs_present = {}, {}, set()

                        def pair_counts(df_pairs: pd.DataFrame) -> Dict[tuple, int]:
                            if df_pairs is None or df_pairs.empty:
                                return {}
                            df_pairs = df_pairs.copy()
                            for col in ("Freq_Pre", "Freq_Post"):
                                if col not in df_pairs.columns:
                                    df_pairs[col] = "<empty>"
                                df_pairs[col] = df_pairs[col].fillna("").replace("", "<empty>")
                            return df_pairs.groupby(["Freq_Pre", "Freq_Post"]).size().astype(int).to_dict()

                        new_by_pair = pair_counts(new_df)
                        miss_by_pair = pair_counts(miss_df)

                        neutral_pairs = {(f, f) for f in (set(pre_counts.keys()) | set(post_counts.keys()))}
                        all_pairs = set(params_by_pair) | set(freq_by_pair) | set(unknown_by_pair) | set(new_by_pair) | set(miss_by_pair) | neutral_pairs | pairs_present

                        for (fpre, fpost) in sorted(all_pairs, key=lambda t: (t[0], t[1])):
                            detailed_rows.append({
                                "Table": name,
                                "KeyColumns": ", ".join(meta.get("key_cols", [])),
                                "FreqColumn": meta.get("freq_col", "N/A"),
                                "Freq_Pre": fpre,
                                "Freq_Post": fpost,
                                "Relations_Pre": int(pre_counts.get(fpre, 0)),
                                "Relations_Post": int(post_counts.get(fpost, 0)),
                                "Parameters_Discrepancies": int(params_by_pair.get((fpre, fpost), 0)),
                                "Freq_Discrepancies": int(freq_by_pair.get((fpre, fpost), 0)),
                                "SSB_Unknown": int(unknown_by_pair.get((fpre, fpost), 0)),
                                "New_Relations": int(new_by_pair.get((fpre, fpost), 0)),
                                "Missing_Relations": int(miss_by_pair.get((fpre, fpost), 0)),
                            })

                detailed_df = pd.DataFrame(detailed_rows) if detailed_rows else pd.DataFrame(columns=["Table", "KeyColumns", "FreqColumn", "Freq_Pre", "Freq_Post", "Relations_Pre", "Relations_Post", "Parameters_Discrepancies", "Freq_Discrepancies", "SSB_Unknown", "New_Relations", "Missing_Relations"])
                detailed_df.to_excel(writer, sheet_name="Summary_CellRelation", index=False)
                written_sheet_dfs["Summary_CellRelation"] = detailed_df

                # Highlight rows where Freq_Pre or Freq_Post matches old or new SSB frequency (Pre/Post) - openpyxl only
                if excel_engine == "openpyxl":
                    try:
                        from openpyxl.styles import PatternFill

                        ws_main = writer.sheets.get("Summary_CellRelation")
                        if ws_main is not None:
                            headers = {ws_main.cell(row=1, column=c).value: c for c in range(1, ws_main.max_column + 1)}
                            col_freq_pre = headers.get("Freq_Pre")
                            col_freq_post = headers.get("Freq_Post")
                            if col_freq_pre and col_freq_post:
                                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                                def _to_int_or_none(v):
                                    try:
                                        if v is None:
                                            return None
                                        s = str(v).strip()
                                        if not s or s == "<empty>":
                                            return None
                                        return int(float(s))
                                    except Exception:
                                        return None

                                old_freq = _to_int_or_none(self.n77_ssb_pre)
                                new_freq = _to_int_or_none(self.n77_ssb_post)
                                target_freqs = {f for f in [old_freq, new_freq] if f is not None}

                                for r in range(2, ws_main.max_row + 1):
                                    vpre = _to_int_or_none(ws_main.cell(row=r, column=col_freq_pre).value)
                                    vpost = _to_int_or_none(ws_main.cell(row=r, column=col_freq_post).value)
                                    if (vpre in target_freqs) or (vpost in target_freqs):
                                        for c in range(1, ws_main.max_column + 1):
                                            ws_main.cell(row=r, column=c).fill = yellow_fill
                    except Exception:
                        pass

                # Collect all dataframes that contain Correction_Cmd to export them later
                correction_cmd_sources: Dict[str, pd.DataFrame] = {}

                # GU sheets
                if results and "GUtranCellRelation" in results:
                    b = results["GUtranCellRelation"]
                    gu_rel_df = b.get("all_relations", pd.DataFrame())
                    gu_disc_df = enforce_gu_columns(b.get("discrepancies"))
                    gu_missing_df = enforce_gu_columns(b.get("missing_in_post"))
                    gu_new_df = enforce_gu_columns(b.get("new_in_post"))

                    freq_only_text = "SSB Post-Retuning keeps equal than SSB Pre-Retuning"
                    unknown_text = "SSB Target Unknown"

                    diff_txt = gu_disc_df["DiffColumns"].astype(str).str.strip() if "DiffColumns" in gu_disc_df.columns else pd.Series("", index=gu_disc_df.index)
                    tgt_txt = gu_disc_df["GNodeB_SSB_Target"].astype(str).str.strip() if "GNodeB_SSB_Target" in gu_disc_df.columns else pd.Series("", index=gu_disc_df.index)

                    mask_unknown = diff_txt.eq(unknown_text) | tgt_txt.eq("Unknown")
                    mask_freq_only = diff_txt.eq(freq_only_text) & tgt_txt.eq("SSB-Post") & (~tgt_txt.eq("Unknown"))
                    mask_param_only = (~mask_freq_only) & (~mask_unknown)

                    gu_freq_disc_df = gu_disc_df[mask_freq_only].copy()
                    gu_param_disc_df = gu_disc_df[mask_param_only].copy()
                    gu_unknown_df = gu_disc_df[mask_unknown].copy()

                    # Build correction commands using external helpers (relations as main source)
                    gu_new_df = build_correction_command_gu_new_relations(gu_new_df, gu_rel_df)
                    gu_missing_df = build_correction_command_gu_missing_relations(gu_missing_df, gu_rel_df, self.n77_ssb_pre, self.n77_ssb_post)
                    gu_freq_disc_cmd_df = build_correction_command_gu_discrepancies(gu_freq_disc_df, gu_rel_df, self.n77_ssb_pre, self.n77_ssb_post)
                    gu_param_disc_cmd_df = build_correction_command_gu_discrepancies(gu_param_disc_df, gu_rel_df, self.n77_ssb_pre, self.n77_ssb_post)

                    correction_cmd_sources["GU_missing"] = gu_missing_df
                    correction_cmd_sources["GU_new"] = gu_new_df
                    correction_cmd_sources["GU_freq_disc"] = gu_freq_disc_cmd_df
                    correction_cmd_sources["GU_param_disc"] = gu_param_disc_cmd_df

                    gu_rel_df.to_excel(writer, sheet_name="GU_relations", index=False);
                    written_sheet_dfs["GU_relations"] = gu_rel_df
                    gu_param_disc_cmd_df.to_excel(writer, sheet_name="GU_param_disc", index=False);
                    written_sheet_dfs["GU_param_disc"] = gu_param_disc_cmd_df
                    gu_freq_disc_cmd_df.to_excel(writer, sheet_name="GU_freq_disc", index=False);
                    written_sheet_dfs["GU_freq_disc"] = gu_freq_disc_cmd_df
                    gu_unknown_df.to_excel(writer, sheet_name="GU_unknown", index=False);
                    written_sheet_dfs["GU_unknown"] = gu_unknown_df
                    gu_missing_df.to_excel(writer, sheet_name="GU_missing", index=False);
                    written_sheet_dfs["GU_missing"] = gu_missing_df
                    gu_new_df.to_excel(writer, sheet_name="GU_new", index=False);
                    written_sheet_dfs["GU_new"] = gu_new_df
                else:
                    df0 = pd.DataFrame();
                    df0.to_excel(writer, sheet_name="GU_relations", index=False);
                    written_sheet_dfs["GU_relations"] = df0
                    df1 = enforce_gu_columns(pd.DataFrame());
                    df1.to_excel(writer, sheet_name="GU_param_disc", index=False);
                    written_sheet_dfs["GU_param_disc"] = df1
                    df2 = enforce_gu_columns(pd.DataFrame());
                    df2.to_excel(writer, sheet_name="GU_freq_disc", index=False);
                    written_sheet_dfs["GU_freq_disc"] = df2
                    df3 = enforce_gu_columns(pd.DataFrame());
                    df3.to_excel(writer, sheet_name="GU_unknown", index=False);
                    written_sheet_dfs["GU_unknown"] = df3
                    df4 = enforce_gu_columns(pd.DataFrame());
                    df4.to_excel(writer, sheet_name="GU_missing", index=False);
                    written_sheet_dfs["GU_missing"] = df4
                    df5 = enforce_gu_columns(pd.DataFrame());
                    df5.to_excel(writer, sheet_name="GU_new", index=False);
                    written_sheet_dfs["GU_new"] = df5

                # NR sheets
                if results and "NRCellRelation" in results:
                    b = results["NRCellRelation"]
                    nr_rel_df = b.get("all_relations", pd.DataFrame())
                    nr_disc_df = enforce_nr_columns(b.get("discrepancies"))
                    nr_missing_df = enforce_nr_columns(b.get("missing_in_post"))
                    nr_new_df = enforce_nr_columns(b.get("new_in_post"))

                    freq_only_text = "SSB Post-Retuning keeps equal than SSB Pre-Retuning"
                    unknown_text = "SSB Target Unknown"

                    diff_txt = nr_disc_df["DiffColumns"].astype(str).str.strip() if "DiffColumns" in nr_disc_df.columns else pd.Series("", index=nr_disc_df.index)
                    tgt_txt = nr_disc_df["GNodeB_SSB_Target"].astype(str).str.strip() if "GNodeB_SSB_Target" in nr_disc_df.columns else pd.Series("", index=nr_disc_df.index)

                    mask_unknown = diff_txt.eq(unknown_text) | tgt_txt.eq("Unknown")
                    mask_freq_only = diff_txt.eq(freq_only_text) & tgt_txt.eq("SSB-Post") & (~tgt_txt.eq("Unknown"))
                    mask_param_only = (~mask_freq_only) & (~mask_unknown)

                    nr_freq_disc_df = nr_disc_df[mask_freq_only].copy()
                    nr_param_disc_df = nr_disc_df[mask_param_only].copy()
                    nr_unknown_df = nr_disc_df[mask_unknown].copy()

                    nr_new_df = build_correction_command_nr_new_relations(nr_new_df, nr_rel_df)
                    nr_missing_df = build_correction_command_nr_missing_relations(nr_missing_df, nr_rel_df, self.n77_ssb_pre, self.n77_ssb_post)
                    nr_freq_disc_cmd_df = build_correction_command_nr_discrepancies(nr_freq_disc_df, nr_rel_df, self.n77_ssb_pre, self.n77_ssb_post)
                    nr_param_disc_cmd_df = build_correction_command_nr_discrepancies(nr_param_disc_df, nr_rel_df, self.n77_ssb_pre, self.n77_ssb_post)

                    correction_cmd_sources["NR_missing"] = nr_missing_df
                    correction_cmd_sources["NR_new"] = nr_new_df
                    correction_cmd_sources["NR_freq_disc"] = nr_freq_disc_cmd_df
                    correction_cmd_sources["NR_param_disc"] = nr_param_disc_cmd_df

                    nr_rel_df.to_excel(writer, sheet_name="NR_relations", index=False);
                    written_sheet_dfs["NR_relations"] = nr_rel_df
                    nr_param_disc_cmd_df.to_excel(writer, sheet_name="NR_param_disc", index=False);
                    written_sheet_dfs["NR_param_disc"] = nr_param_disc_cmd_df
                    nr_freq_disc_cmd_df.to_excel(writer, sheet_name="NR_freq_disc", index=False);
                    written_sheet_dfs["NR_freq_disc"] = nr_freq_disc_cmd_df
                    nr_unknown_df.to_excel(writer, sheet_name="NR_unknown", index=False);
                    written_sheet_dfs["NR_unknown"] = nr_unknown_df
                    nr_missing_df.to_excel(writer, sheet_name="NR_missing", index=False);
                    written_sheet_dfs["NR_missing"] = nr_missing_df
                    nr_new_df.to_excel(writer, sheet_name="NR_new", index=False);
                    written_sheet_dfs["NR_new"] = nr_new_df
                else:
                    df0 = pd.DataFrame();
                    df0.to_excel(writer, sheet_name="NR_relations", index=False);
                    written_sheet_dfs["NR_relations"] = df0
                    df1 = enforce_nr_columns(pd.DataFrame());
                    df1.to_excel(writer, sheet_name="NR_param_disc", index=False);
                    written_sheet_dfs["NR_param_disc"] = df1
                    df2 = enforce_nr_columns(pd.DataFrame());
                    df2.to_excel(writer, sheet_name="NR_freq_disc", index=False);
                    written_sheet_dfs["NR_freq_disc"] = df2
                    df3 = enforce_nr_columns(pd.DataFrame());
                    df3.to_excel(writer, sheet_name="NR_unknown", index=False);
                    written_sheet_dfs["NR_unknown"] = df3
                    df4 = enforce_nr_columns(pd.DataFrame());
                    df4.to_excel(writer, sheet_name="NR_missing", index=False);
                    written_sheet_dfs["NR_missing"] = df4
                    df5 = enforce_nr_columns(pd.DataFrame());
                    df5.to_excel(writer, sheet_name="NR_new", index=False);
                    written_sheet_dfs["NR_new"] = df5

                # Export text files (outside GU/NR blocks)
                if correction_cmd_sources:
                    cmd_files = export_relations_commands(output_dir, correction_cmd_sources, base_folder_name="Correction_Cmd_CC", module_name=module_name, market_tag=market_tag)

                # -------------------------------------------------------------------
                #  APPLY HEADER STYLING + AUTO-FIT COLUMNS FOR ALL SHEETS
                # -------------------------------------------------------------------
                if excel_engine == "xlsxwriter":
                    style_headers_autofilter_and_autofit_xlsxwriter(writer, sheet_dfs=written_sheet_dfs, freeze_header=True, align="left", max_autofit_rows=fast_excel_autofit_rows, max_col_width=fast_excel_autofit_max_width, enable_a1_hyperlinks=True, hyperlink_sheet="Summary_CellRelation")

                    # Hyperlink A1 -> Summary_CellRelation (xlsxwriter)
                    try:
                        link_fmt = writer.book.add_format({"font_color": "#0563C1", "underline": 1})
                        for ws_name, ws in writer.sheets.items():
                            if ws_name == "Summary_CellRelation":
                                continue
                            df_for_ws = written_sheet_dfs.get(ws_name)
                            header_text = str(list(df_for_ws.columns)[0]) if df_for_ws is not None and hasattr(df_for_ws, "columns") and len(list(df_for_ws.columns)) > 0 else "A1"
                            ws.write_url(0, 0, "internal:Summary_CellRelation!A1", link_fmt, header_text)
                    except Exception:
                        pass
                else:
                    color_summary_tabs(writer, prefix="Summary", rgb_hex="00B050")
                    style_headers_autofilter_and_autofit(writer, freeze_header=True, align="left", enable_a1_hyperlink=True, hyperlink_sheet="Summary_CellRelation")

                    ws_comp = writer.sheets.get("SummaryAuditComparisson")
                    if ws_comp is not None:
                        apply_alternating_category_row_fills(ws_comp, value_header="Value_Post")


            _move_into_place(tmp_excel_cc_long, excel_cc_cell_relation_long)

            # -------------------------------------------------------------------
            #  Write CellRelations.xlsx (TEMP)
            # -------------------------------------------------------------------
            print(f"{module_name} {market_tag} [INFO] Saving {pretty_path(excel_cell_relation)} (engine={excel_engine}, temp -> final)...")

            written_sheet_dfs2: Dict[str, pd.DataFrame] = {}

            if excel_engine == "xlsxwriter":
                writer_ctx2 = pd.ExcelWriter(tmp_excel_cell_long, engine="xlsxwriter", engine_kwargs={"options": {"strings_to_urls": False, "strings_to_numbers": False}})
            else:
                writer_ctx2 = pd.ExcelWriter(tmp_excel_cell_long, engine="openpyxl")

            with writer_ctx2 as writer:
                if "GUtranCellRelation" in self.tables:
                    df_gu = self.tables["GUtranCellRelation"]
                    df_gu.to_excel(writer, sheet_name="GU_all", index=False)
                    written_sheet_dfs2["GU_all"] = df_gu
                if "NRCellRelation" in self.tables:
                    df_nr = self.tables["NRCellRelation"]
                    df_nr.to_excel(writer, sheet_name="NR_all", index=False)
                    written_sheet_dfs2["NR_all"] = df_nr

                # -------------------------------------------------------------------
                #  APPLY HEADER STYLING + AUTO-FIT COLUMNS FOR ALL SHEETS
                # -------------------------------------------------------------------
                if excel_engine == "xlsxwriter":
                    style_headers_autofilter_and_autofit_xlsxwriter(writer, sheet_dfs=written_sheet_dfs2, freeze_header=True, align="left", max_autofit_rows=fast_excel_autofit_rows, max_col_width=fast_excel_autofit_max_width, enable_a1_hyperlinks=False, hyperlink_sheet="")
                else:
                    style_headers_autofilter_and_autofit(writer, freeze_header=True, align="left")

            _move_into_place(tmp_excel_cell_long, excel_cell_relation_long)

        finally:
            try:
                if tmp_dir_cc and os.path.isdir(tmp_dir_cc):
                    shutil.rmtree(tmp_dir_cc, ignore_errors=True)
            except Exception:
                pass

            try:
                if tmp_dir_cell and os.path.isdir(tmp_dir_cell):
                    shutil.rmtree(tmp_dir_cell, ignore_errors=True)
            except Exception:
                pass

